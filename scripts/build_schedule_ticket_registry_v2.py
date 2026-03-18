#!/usr/bin/env python3
import json
from collections import Counter, defaultdict
from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/macbook15/Downloads/MacAi")
ARTIFACTS = ROOT / "artifacts"

V1_PATH = ARTIFACTS / "schedule_2026-03-10_ticket_registry_v1.json"
DICT_V2_PATH = ARTIFACTS / "ticket_dictionary_v2.json"
PROBE_V2_PATH = ARTIFACTS / "schedule_2026-03-10_ticket_semantic_probe_v2.json"

OUT_JSON = ARTIFACTS / "schedule_2026-03-10_ticket_registry_v2.json"
OUT_XLSX = ARTIFACTS / "schedule_2026-03-10_ticket_registry_v2.xlsx"
OUT_MD = ARTIFACTS / "schedule_2026-03-10_ticket_registry_v2.md"


def read_json(path: Path):
    return json.loads(path.read_text())


def write_json(path: Path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2))


def freeze(v):
    if isinstance(v, dict):
        return tuple(sorted((k, freeze(val)) for k, val in v.items()))
    if isinstance(v, list):
        return tuple(freeze(x) for x in v)
    return v


def ordered_unique(seq):
    seen = set()
    out = []
    for item in seq:
        marker = freeze(item)
        if marker not in seen:
            seen.add(marker)
            out.append(item)
    return out


def reuse_safety(category: str, proof_level: str) -> str:
    if category == "coordinator_person_reference_marker" or proof_level == "inferred":
        return "no"
    if category in {"external_booking_source_info", "free_text_note_marker"}:
        return "conditional"
    return "yes"


def enrich_dictionary_row(row: dict) -> dict:
    item = deepcopy(row)
    icon_class = item["icon_class"]
    tooltip = (item["tooltip_text"] or "").strip()
    category = item.get("normalized_category_v2") or item.get("normalized_category") or ""
    proof = item.get("normalized_category_proof_v2") or item.get("proof_level") or ""

    if icon_class == "schi schi-10":
        category = "coordinator_person_reference_marker"
        proof = "inferred"
        proof_scope = "observed_family_2026_03_10_schedule_subset"
        evidence_basis = (
            "live_rendered_rc_tooltip_person_name + focused_runtime_probe_same_name_in_coordinator_context"
        )
        note = (
            "Icon-local tooltip contains only a person name. Same names are present in runtime DOM under "
            "coordinator-linked sections, but icon-local role label was not found."
        )
    elif icon_class == "schi schi-3" and tooltip.startswith("ПроДокторов."):
        category = "external_booking_source_info"
        proof = "proven"
        proof_scope = "exact_tooltip_text_rule"
        evidence_basis = "live_rendered_rc_tooltip_exact_text"
        note = "Tooltip explicitly identifies external booking source and patient booking context."
    elif icon_class == "schi schi-3":
        category = "free_text_note_marker"
        proof = "proven"
        proof_scope = "observed_non_external_schi_3_cases_2026_03_10"
        evidence_basis = "live_rendered_rc_tooltip_exact_text"
        note = "Tooltip is a plain free-text note; note author/owner is not proven."
    else:
        if proof.startswith("proven"):
            proof = "proven"
        proof_scope = "exact_tooltip_text_rule" if proof == "proven" else "observed_family_2026_03_10_schedule_subset"
        evidence_basis = (
            "live_rendered_rc_tooltip_exact_text"
            if proof == "proven"
            else "focused_schedule_runtime_probe"
        )
        note = item.get("semantic_note_v2", "") or item.get("notes", "")

    item["normalized_category_v2"] = category
    item["proof_level"] = proof
    item["proof_scope"] = proof_scope
    item["evidence_basis"] = evidence_basis
    item["safe_for_global_reuse"] = reuse_safety(category, proof)
    item["semantic_note_v2"] = note
    item.pop("normalized_category_proof_v2", None)
    return item


def enrich_ticket_row(row: dict, dictionary_index: dict) -> dict:
    item = deepcopy(row)
    key = (item["icon_class"], (item.get("tooltip_text") or "").strip())
    d = dictionary_index[key]
    item["normalized_category_v2"] = d["normalized_category_v2"]
    item["proof_level_v2"] = d["proof_level"]
    item["proof_scope_v2"] = d["proof_scope"]
    item["evidence_basis_v2"] = d["evidence_basis"]
    item["safe_for_global_reuse_v2"] = d["safe_for_global_reuse"]
    item["semantic_note_v2"] = d["semantic_note_v2"]
    return item


def aggregate_visit(row: dict, ticket_rows: list[dict]) -> dict:
    item = deepcopy(row)
    categories = ordered_unique([t["normalized_category_v2"] for t in ticket_rows if t["normalized_category_v2"]])
    proof_levels = ordered_unique([t["proof_level_v2"] for t in ticket_rows if t["proof_level_v2"]])
    evidence_bases = ordered_unique([t["evidence_basis_v2"] for t in ticket_rows if t["evidence_basis_v2"]])
    reuse_values = [t["safe_for_global_reuse_v2"] for t in ticket_rows if t["safe_for_global_reuse_v2"]]
    if "no" in reuse_values:
        reuse = "no"
    elif "conditional" in reuse_values:
        reuse = "conditional"
    else:
        reuse = "yes"

    coord_names = ordered_unique(
        [t["tooltip_text"] for t in ticket_rows if t["normalized_category_v2"] == "coordinator_person_reference_marker"]
    )
    note_values = ordered_unique(
        [t["tooltip_text"] for t in ticket_rows if t["normalized_category_v2"] == "free_text_note_marker"]
    )
    external_sources = ordered_unique(
        [t["tooltip_text"] for t in ticket_rows if t["normalized_category_v2"] == "external_booking_source_info"]
    )

    item["ticket_semantics_v2"] = "; ".join(categories)
    item["ticket_semantics_proof_levels_v2"] = "; ".join(proof_levels)
    item["ticket_semantics_evidence_basis_v2"] = "; ".join(evidence_bases)
    item["ticket_semantics_safe_for_global_reuse_v2"] = reuse
    item["ticket_semantics_has_inferred_v2"] = "yes" if "inferred" in proof_levels else "no"
    item["coordinator_person_reference_names_v2"] = "; ".join(coord_names)
    item["free_text_note_values_v2"] = "; ".join(note_values)
    item["external_booking_sources_v2"] = "; ".join(external_sources)
    return item


def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 60)


def add_sheet_from_dicts(wb: Workbook, title: str, rows: list[dict]):
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["empty"])
        return ws
    headers = list(rows[0].keys())
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append([json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v for v in row.values()])
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def main():
    v1 = read_json(V1_PATH)
    existing_dict = read_json(DICT_V2_PATH)
    probe = read_json(PROBE_V2_PATH)

    # Keep a compact evidence appendix from the focused probe for schi-10 and disputed schi-3.
    probe_results = []
    for r in probe["results"]:
        probe_results.append(
            {
                "visit_id": r["visit_id"],
                "icon_class": r["icon_class"],
                "tooltip_text_runtime": r.get("tooltip_text_runtime", ""),
                "patient_name_full": r.get("patient_name_full", ""),
                "doctor_name": r.get("doctor_name", ""),
                "contains_role_term_in_popup": r.get("popup_role_term_detected", False),
                "page_role_term_match_count": len(r.get("page_role_term_matches", [])),
                "final_probe_note": (
                    "No direct role term found in icon-local tooltip/popup"
                    if r["icon_class"] == "schi schi-10"
                    else "Free-text tooltip confirmed by focused runtime probe"
                ),
            }
        )

    enriched_dict = [enrich_dictionary_row(row) for row in existing_dict]
    dictionary_index = {(r["icon_class"], (r["tooltip_text"] or "").strip()): r for r in enriched_dict}

    tickets_flat_v2 = [enrich_ticket_row(row, dictionary_index) for row in v1["tickets_flat"]]
    tickets_by_visit = defaultdict(list)
    for row in tickets_flat_v2:
        tickets_by_visit[row["visit_id"]].append(row)

    visits_registry_v2 = [aggregate_visit(row, tickets_by_visit[row["visit_id"]]) for row in v1["visits_registry"]]

    counter_categories = Counter(t["normalized_category_v2"] for t in tickets_flat_v2)
    counter_proof = Counter(t["proof_level_v2"] for t in tickets_flat_v2)
    qc = {
        "visits_total": len(visits_registry_v2),
        "tickets_total": len(tickets_flat_v2),
        "tickets_with_proven_semantics_v2": counter_proof.get("proven", 0),
        "tickets_with_inferred_semantics_v2": counter_proof.get("inferred", 0),
        "visits_with_any_inferred_ticket_semantics_v2": sum(
            1 for row in visits_registry_v2 if row["ticket_semantics_has_inferred_v2"] == "yes"
        ),
        "schi_10_ticket_count": counter_categories.get("coordinator_person_reference_marker", 0),
        "schi_3_external_source_count": counter_categories.get("external_booking_source_info", 0),
        "schi_3_free_text_count": counter_categories.get("free_text_note_marker", 0),
    }

    out = {
        "target_date": v1["target_date"],
        "subset_rule": v1["subset_rule"],
        "registry_version": "v2",
        "delta_basis": (
            "ticket_registry_v1 + focused semantic probe for schi-10 and disputed schi-3; "
            "no broad rerun of the full schedule screen"
        ),
        "quality_control": qc,
        "visits_registry": visits_registry_v2,
        "tickets_flat": tickets_flat_v2,
        "ticket_dictionary": enriched_dict,
        "focused_probe_appendix": probe_results,
    }

    write_json(OUT_JSON, out)
    write_json(DICT_V2_PATH, enriched_dict)

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet_from_dicts(wb, "summary", [{"metric": k, "value": v} for k, v in qc.items()])
    add_sheet_from_dicts(wb, "visits_registry_v2", visits_registry_v2)
    add_sheet_from_dicts(wb, "tickets_flat_v2", tickets_flat_v2)
    add_sheet_from_dicts(wb, "ticket_dictionary_v2", enriched_dict)
    add_sheet_from_dicts(wb, "focused_probe_v2", probe_results)
    wb.save(OUT_XLSX)

    md = [
        "# Schedule Ticket Registry V2",
        "",
        f"- target_date: `{v1['target_date']}`",
        "- basis: `ticket_registry_v1 + schedule_2026-03-10_ticket_semantic_probe_v2.json`",
        "- broad rerun: `no`",
        "",
        "## Quality Control",
        "",
        "| Metric | Value |",
        "|---|---:|",
    ]
    for k, v in qc.items():
        md.append(f"| {k} | {v} |")
    md.extend(
        [
            "",
            "## Final Semantic Closure",
            "",
            "- `schi-10` -> `coordinator_person_reference_marker`, proof=`inferred`, safe_for_global_reuse=`no`.",
            "- `schi-3` with exact `ПроДокторов.` tooltip -> `external_booking_source_info`, proof=`proven`, safe_for_global_reuse=`conditional`.",
            "- `schi-3` non-external observed cases -> `free_text_note_marker`, proof=`proven`, safe_for_global_reuse=`conditional`.",
            "",
            "## Visit-Level Propagation",
            "",
            "- `visits_registry_v2` now contains aggregated semantic fields from `ticket_dictionary_v2`.",
            "- Downstream consumers no longer need to interpret ticket semantics from raw `icon_class` alone.",
        ]
    )
    OUT_MD.write_text("\n".join(md))


if __name__ == "__main__":
    main()
