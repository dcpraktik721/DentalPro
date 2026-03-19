#!/usr/bin/env python3
"""Build normalized workbook and meta sidecar for RPT_4 from saved repo artifacts."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path("/Users/macbook15/Downloads/MacAi/DentalPro")
REPORTING_DIR = REPO_ROOT / "reporting"
REPORTS_DIR = REPO_ROOT / "reports"
EXCEL_DIR = REPO_ROOT / "excel"
ARTIFACTS_DIR = REPO_ROOT / "artifacts"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EFEFEF")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
BOLD = Font(bold=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text())


def stringify(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(str(cell.value))), 80)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width + 2, 12)


def write_table(ws, start_row: int, rows: list[dict[str, Any]], title: str | None = None) -> int:
    current_row = start_row
    if title:
        ws.cell(current_row, 1, title).font = BOLD
        ws.cell(current_row, 1).fill = SECTION_FILL
        current_row += 1
    if not rows:
        ws.cell(current_row, 1, "no_rows")
        return current_row + 2
    headers: list[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
    for row_idx, row in enumerate(rows, start=current_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, stringify(row.get(header)))
    return current_row + len(rows) + 2


def parse_money(value: str | None) -> float | None:
    if value is None:
      return None
    normalized = str(value).replace("₽", "").replace(" ", "").replace(",", ".").strip()
    if not normalized:
      return None
    try:
      return float(normalized)
    except ValueError:
      return None


def build_summary_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    verification = payload["verification"]
    return [
        {"section": "run", "metric": "report_code", "value": "RPT_4", "notes": ""},
        {"section": "run", "metric": "report_name", "value": "Первичные пациенты", "notes": ""},
        {"section": "run", "metric": "date_start", "value": payload["period"]["start"], "notes": ""},
        {"section": "run", "metric": "date_end", "value": payload["period"]["end"], "notes": ""},
        {"section": "run", "metric": "generated_at", "value": payload["generated_at"], "notes": ""},
        {"section": "runtime", "metric": "ui_report_url", "value": payload["ui"]["report_url"], "notes": ""},
        {"section": "runtime", "metric": "ui_row_count", "value": payload["ui"]["row_count"], "notes": ""},
        {"section": "api", "metric": "api_row_count", "value": len(payload["api_rows"]), "notes": ""},
        {"section": "verification", "metric": "verification_pass", "value": verification["pass"], "notes": ""},
        {"section": "verification", "metric": "verified_fields", "value": verification["verified_fields"], "notes": ""},
        {"section": "verification", "metric": "mismatch_count", "value": len(verification["mismatches"]), "notes": ""},
        {"section": "boundary", "metric": "safe_claim_boundary", "value": "UI-anchored API replay for this date. Row identity and selected field semantics are verified against the live report; hidden backend predicate plus Филиал/Координатор remain unclosed.", "notes": ""},
    ]


def build_normalized_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for idx, row in enumerate(payload["api_rows"], start=1):
        rows.append({
            "report_code": "RPT_4",
            "report_name": "Первичные пациенты",
            "date_start": payload["period"]["start"],
            "date_end": payload["period"]["end"],
            "row_source": "ui_anchored_api_replay",
            "row_index": idx,
            "patient_name": row.get("Пациент"),
            "card_number": row.get("Номер карты"),
            "birth_date": row.get("Дата рождения"),
            "first_record_date": row.get("Дата первой записи"),
            "current_primary_date": row.get("Дата текущей первичной записи"),
            "spent_total": parse_money(row.get("Сумма, потраченная пациентом в клинике")),
            "balance": parse_money(row.get("Актуальный баланс")),
            "source_name": row.get("Источник"),
            "doctor_name": row.get("Врач"),
            "branch_status": "not_fully_verified",
            "coordinator_status": "not_fully_verified",
            "notes": "",
        })
    return rows


def build_warning_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    mismatches = payload["verification"]["mismatches"]
    if mismatches:
        warnings.append({
            "warning_code": "VERIFIED_FIELD_MISMATCH",
            "severity": "high",
            "layer": "ui_vs_api",
            "message": f"Verified fields mismatched on {len(mismatches)} comparisons.",
        })
    warnings.extend([
        {
            "warning_code": "HIDDEN_BACKEND_PREDICATE",
            "severity": "high",
            "layer": "selection",
            "message": "Full row selection logic for RPT_4 is still not closed; at least one hidden backend predicate remains.",
        },
        {
            "warning_code": "BRANCH_COORDINATOR_NOT_CLOSED",
            "severity": "medium",
            "layer": "field_closure",
            "message": "Филиал and Координатор remain intentionally unclosed in the accepted API replay.",
        },
    ])
    return warnings


def build_notes_rows(date: str) -> list[dict[str, Any]]:
    return [
        {"field": "live_report_md", "value": (REPORTS_DIR / f"RPT_4_live_report_{date}.md").read_text()},
        {"field": "problem_log_md", "value": (REPORTS_DIR / f"RPT_4_problem_log_{date}.md").read_text()},
    ]


def build_meta(payload: dict[str, Any], output_path: Path) -> dict[str, Any]:
    verification = payload["verification"]
    return {
        "report_code": "RPT_4",
        "report_name": "Первичные пациенты",
        "date_start": payload["period"]["start"],
        "date_end": payload["period"]["end"],
        "generated_at": payload["generated_at"],
        "run_status": "success" if verification["pass"] else "success_with_warnings",
        "primary_truth": "runtime_html_anchored_api_replay",
        "runtime_checked": True,
        "export_checked": False,
        "export_available": False,
        "runtime_row_count": payload["ui"]["row_count"],
        "export_row_count": None,
        "parity_status": "verified_fields_pass" if verification["pass"] else "verified_fields_mismatch",
        "warnings": build_warning_rows(payload),
        "safe_claim_boundary": "UI-anchored API replay for this date. Row identity and selected field semantics are verified against the live report; hidden backend predicate plus Филиал/Координатор remain unclosed.",
        "normalized_workbook_path": str(output_path),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default="2026-02-18")
    parser.add_argument("--output")
    args = parser.parse_args()

    date = args.date
    payload = read_json(REPORTING_DIR / f"report-4-api-copy-{date}-{date}.json")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "RPT_4_summary"
    write_table(ws_summary, 1, build_summary_rows(payload), title="summary")

    ws_norm = wb.create_sheet("RPT_4_normalized")
    write_table(ws_norm, 1, build_normalized_rows(payload), title="normalized_rows")

    ws_warn = wb.create_sheet("RPT_4_warnings")
    end_row = write_table(ws_warn, 1, build_warning_rows(payload), title="warnings")
    for row in ws_warn.iter_rows(min_row=3, max_row=end_row):
        for cell in row:
            if cell.value is not None:
                cell.fill = WARN_FILL

    ws_notes = wb.create_sheet("RPT_4_notes")
    write_table(ws_notes, 1, build_notes_rows(date), title="notes")

    for ws in wb.worksheets:
        autofit(ws)

    output_path = Path(args.output) if args.output else EXCEL_DIR / f"RPT_4_{date}_normalized.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    load_workbook(output_path)

    meta_path = ARTIFACTS_DIR / f"RPT_4_{date}_meta.json"
    meta_path.write_text(json.dumps(build_meta(payload, output_path), ensure_ascii=False, indent=2) + "\n")
    print(output_path)
    print(meta_path)


if __name__ == "__main__":
    main()
