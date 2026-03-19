#!/usr/bin/env python3
"""Build normalized workbook and meta sidecar for RPT_18 from a saved verification artifact."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path("/Users/macbook15/Downloads/MacAi/DentalPro")
REPORTING_DIR = REPO_ROOT / "reporting"
REPORTS_DIR = REPO_ROOT / "reports"
EXCEL_DIR = REPO_ROOT / "excel"
ARTIFACTS_DIR = REPO_ROOT / "artifacts"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EFEFEF")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
BOLD = Font(bold=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text())


def stringify(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width = len(str(cell.value))
            widths[cell.column] = min(max(widths.get(cell.column, 0), width), 80)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width + 2, 12)


def write_table(ws, start_row: int, rows: list[dict[str, Any]], title: str | None = None) -> int:
    current_row = start_row
    if title:
        ws.cell(current_row, 1, title).font = BOLD
        ws.cell(current_row, 1).fill = SECTION_FILL
        current_row += 1

    if not rows:
        ws.cell(current_row, 1, "no_rows")
        return current_row + 2

    headers: list[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    for row_idx, row in enumerate(rows, start=current_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, stringify(row.get(header)))

    return current_row + len(rows) + 2


def build_summary_rows(artifact: dict[str, Any], meta: dict[str, Any]) -> list[dict[str, Any]]:
    verdict = artifact["verdict"]
    ui = artifact["ui"]
    schedule = artifact["schedule_evidence"]
    return [
        {"section": "run", "metric": "report_code", "value": meta["report_code"], "notes": ""},
        {"section": "run", "metric": "report_name", "value": meta["report_name"], "notes": ""},
        {"section": "run", "metric": "date_start", "value": meta["date_start"], "notes": ""},
        {"section": "run", "metric": "date_end", "value": meta["date_end"], "notes": ""},
        {"section": "run", "metric": "run_status", "value": meta["run_status"], "notes": ""},
        {"section": "run", "metric": "generated_at", "value": meta["generated_at"], "notes": ""},
        {"section": "truth", "metric": "primary_truth", "value": meta["primary_truth"], "notes": ""},
        {"section": "truth", "metric": "parity_status", "value": meta["parity_status"], "notes": ""},
        {"section": "runtime", "metric": "runtime_title", "value": ui.get("title"), "notes": ""},
        {"section": "runtime", "metric": "runtime_empty", "value": ui.get("empty"), "notes": ""},
        {"section": "runtime", "metric": "runtime_row_count", "value": meta["runtime_row_count"], "notes": ""},
        {"section": "runtime", "metric": "runtime_total_patients", "value": ui.get("totals", {}).get("patients_count"), "notes": ""},
        {"section": "api", "metric": "api_method", "value": artifact.get("api_method"), "notes": ""},
        {"section": "api", "metric": "api_row_count", "value": meta["api_row_count"], "notes": ""},
        {"section": "api", "metric": "schedule_unique_patients", "value": schedule.get("uniquePatientsCount"), "notes": ""},
        {"section": "verdict", "metric": "overall_pass", "value": verdict.get("pass"), "notes": ""},
        {"section": "verdict", "metric": "row_level_pass", "value": verdict.get("row_level_pass"), "notes": ""},
        {"section": "verdict", "metric": "totals_pass", "value": verdict.get("totals_pass"), "notes": ""},
        {"section": "verdict", "metric": "comparable_totals_pass", "value": verdict.get("comparable_totals_pass"), "notes": ""},
        {"section": "verdict", "metric": "unique_patient_totals_pass", "value": verdict.get("unique_patient_totals_pass"), "notes": ""},
        {"section": "boundary", "metric": "safe_claim_boundary", "value": meta["safe_claim_boundary"], "notes": ""},
    ]


def build_normalized_rows(artifact: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    ui_rows_by_doctor = {row["doctor"]: row for row in artifact["ui"].get("rows", [])}
    schedule = artifact["schedule_evidence"].get("perDoctor", {})
    for idx, row in enumerate(artifact["api"]["rows"], start=1):
        ui_row = ui_rows_by_doctor.get(row["doctor"], {})
        doctor_schedule = schedule.get(row["doctor_id"], {})
        rows.append(
            {
                "report_code": "RPT_18",
                "report_name": "Ключевые показатели эффективности",
                "date_start": artifact["target_date_start"],
                "date_end": artifact["target_date_end"],
                "row_source": "api_composite_validated_against_runtime_html",
                "row_index": idx,
                "doctor_id": row["doctor_id"],
                "doctor_name": row["doctor"],
                "money_per_hour_rub": row["money_per_hour_rub"],
                "planned_hours": row["planned_hours"],
                "actual_hours": row["actual_hours"],
                "load_pct": row["load_pct"],
                "plan_rub": row["plan_rub"],
                "revenue_rub": row["revenue_rub"],
                "complete_pct": row["complete_pct"],
                "primary_count": row["primary_count"],
                "visits_count": row["visits_count"],
                "patients_count": row["patients_count"],
                "visits_per_patient": row["visits_per_patient"],
                "revenue_per_visit": row["revenue_per_visit"],
                "revenue_per_patient": row["revenue_per_patient"],
                "revenue_per_hour": row["revenue_per_hour"],
                "primary_share_pct": row["primary_share_pct"],
                "repeat_pct": row["repeat_pct"],
                "runtime_doctor_match": bool(ui_row),
                "runtime_revenue_rub": ui_row.get("revenue_rub"),
                "schedule_unique_patients_count": doctor_schedule.get("unique_patients_count"),
                "schedule_visit_count": doctor_schedule.get("visit_count"),
                "notes": "dual-source row validated by verifier",
            }
        )
    return rows


def build_totals_rows(artifact: dict[str, Any]) -> list[dict[str, Any]]:
    totals = artifact["comparisons"]["totals"]["details"]
    ui_totals = totals.get("uiTotals", {})
    api_comparable = totals.get("apiComparableTotals", {})
    api_unique = totals.get("apiUniquePatientTotals", {})
    rows = []
    for field in [
        "planned_hours",
        "actual_hours",
        "load_pct",
        "plan_rub",
        "revenue_rub",
        "complete_pct",
        "primary_count",
        "visits_count",
        "patients_count",
        "visits_per_patient",
        "revenue_per_visit",
        "revenue_per_patient",
        "revenue_per_hour",
        "primary_share_pct",
        "repeat_pct",
    ]:
        rows.append(
            {
                "metric": field,
                "ui_total": ui_totals.get(field),
                "api_comparable_total": api_comparable.get(field),
                "api_unique_patient_total": api_unique.get(field),
                "notes": "unique-patient-derived metrics come from mobile/schedule supplement" if field in {"patients_count", "visits_per_patient", "revenue_per_patient", "primary_share_pct"} else "",
            }
        )
    return rows


def build_warning_rows(meta: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "warning_code": item["code"],
            "severity": item["severity"],
            "layer": item["layer"],
            "message": item["message"],
        }
        for item in meta["warnings"]
    ]


def build_notes_rows(report_md: Path, problem_md: Path) -> list[dict[str, Any]]:
    return [
        {"field": "live_report_md", "value": report_md.read_text()},
        {"field": "problem_log_md", "value": problem_md.read_text()},
    ]


def build_meta(artifact: dict[str, Any], workbook_path: Path) -> dict[str, Any]:
    verdict = artifact["verdict"]
    warnings = [
        {
            "code": "DUAL_SOURCE_REQUIRED",
            "severity": "medium",
            "layer": "source_classification",
            "message": "RPT_18 requires mobile/owner/efficiency rows plus mobile/schedule unique-patient supplement; a single accepted API is not sufficient for full totals closure.",
        },
        {
            "code": "FAMILY_SPECIFIC_CANONICALIZATION",
            "severity": "low",
            "layer": "normalization",
            "message": "Grouped-header/business-header projection is proven for the RPT_18 family and must not be promoted automatically to unrelated report families.",
        },
    ]
    return {
        "report_code": "RPT_18",
        "report_name": "Ключевые показатели эффективности",
        "date_start": artifact["target_date_start"],
        "date_end": artifact["target_date_end"],
        "generated_at": artifact["generated_at"],
        "run_status": "success" if verdict["pass"] else "partial",
        "primary_truth": "accepted_api_composite_validated_against_runtime_html",
        "runtime_checked": True,
        "export_checked": False,
        "export_available": False,
        "runtime_row_count": len(artifact["ui"].get("rows", [])),
        "api_row_count": artifact["api"]["row_count"],
        "export_row_count": None,
        "parity_status": "confirmed" if verdict["pass"] else "mismatch_detected",
        "warnings": warnings,
        "safe_claim_boundary": "For 2026-02-18, RPT_18 is verified via the accepted API composite mobile/owner/efficiency plus mobile/schedule for unique-patient totals, validated against runtime UI. This does not prove that a single API endpoint fully reproduces the report, and the normalization policy remains family-specific.",
        "source_artifact_path": str(REPORTING_DIR / f"efficiency-report-verification-{artifact['target_range_label']}.json"),
        "workbook_path": str(workbook_path),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True)
    parser.add_argument("--input")
    parser.add_argument("--output")
    parser.add_argument("--meta-output")
    args = parser.parse_args()

    date = args.date
    input_path = Path(args.input) if args.input else REPORTING_DIR / f"efficiency-report-verification-{date}.json"
    output_path = Path(args.output) if args.output else EXCEL_DIR / f"RPT_18_{date}_normalized.xlsx"
    meta_path = Path(args.meta_output) if args.meta_output else ARTIFACTS_DIR / f"RPT_18_{date}_meta.json"

    artifact = read_json(input_path)
    meta = build_meta(artifact, output_path)

    report_md = REPORTS_DIR / f"RPT_18_live_report_{date}.md"
    problem_md = REPORTS_DIR / f"RPT_18_problem_log_{date}.md"

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "RPT_18_summary"
    write_table(ws_summary, 1, build_summary_rows(artifact, meta), title="summary")

    ws_norm = wb.create_sheet("RPT_18_normalized")
    write_table(ws_norm, 1, build_normalized_rows(artifact), title="normalized_rows")

    ws_totals = wb.create_sheet("RPT_18_totals")
    write_table(ws_totals, 1, build_totals_rows(artifact), title="totals")

    ws_warn = wb.create_sheet("RPT_18_warnings")
    end_row = write_table(ws_warn, 1, build_warning_rows(meta), title="warnings")
    for row in ws_warn.iter_rows(min_row=3, max_row=end_row):
        for cell in row:
            if cell.value is not None:
                cell.fill = WARN_FILL

    ws_notes = wb.create_sheet("RPT_18_notes")
    write_table(ws_notes, 1, build_notes_rows(report_md, problem_md), title="notes")

    for ws in wb.worksheets:
        autofit(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    meta_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    load_workbook(output_path)
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2) + "\n")
    print(output_path)
    print(meta_path)


if __name__ == "__main__":
    main()
