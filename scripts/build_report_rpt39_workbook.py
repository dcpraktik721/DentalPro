#!/usr/bin/env python3
"""Build normalized workbook and meta sidecar for RPT_39 from saved repo artifacts."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path("/Users/macbook15/Downloads/MacAi/DentalPro")
REPORTING_DIR = REPO_ROOT / "reporting"
REPORTS_DIR = REPO_ROOT / "reports"
EXCEL_DIR = REPO_ROOT / "excel"
ARTIFACTS_DIR = REPO_ROOT / "artifacts"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EFEFEF")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
BOLD = Font(bold=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text())


def stringify(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(str(cell.value))), 80)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width + 2, 12)


def write_table(ws, start_row: int, rows: list[dict[str, Any]], title: str | None = None) -> int:
    current_row = start_row
    if title:
        ws.cell(current_row, 1, title).font = BOLD
        ws.cell(current_row, 1).fill = SECTION_FILL
        current_row += 1
    if not rows:
        ws.cell(current_row, 1, "no_rows")
        return current_row + 2
    headers: list[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
    for row_idx, row in enumerate(rows, start=current_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, stringify(row.get(header)))
    return current_row + len(rows) + 2


def build_summary_rows(curated: dict[str, Any]) -> list[dict[str, Any]]:
    auth = curated["auth"]
    runtime = curated["runtime"]
    sc = curated["source_classification"]
    return [
        {"section": "run", "metric": "report_code", "value": curated["report_code"], "notes": ""},
        {"section": "run", "metric": "report_name", "value": curated["report_name"], "notes": ""},
        {"section": "run", "metric": "date_start", "value": curated["date_start"], "notes": ""},
        {"section": "run", "metric": "date_end", "value": curated["date_end"], "notes": ""},
        {"section": "run", "metric": "generated_at", "value": curated["generated_at"], "notes": ""},
        {"section": "run", "metric": "run_status", "value": curated["run_status"], "notes": ""},
        {"section": "auth", "metric": "authenticated", "value": auth["authenticated"], "notes": ""},
        {"section": "auth", "metric": "final_url", "value": auth["final_url"], "notes": ""},
        {"section": "auth", "metric": "redirect_detected", "value": auth["redirect_detected"], "notes": ""},
        {"section": "runtime", "metric": "primary_truth", "value": sc["primary_truth"], "notes": ""},
        {"section": "runtime", "metric": "table_count", "value": runtime["structure"]["table_count"], "notes": ""},
        {"section": "runtime", "metric": "distinct_leaf_rows_observed", "value": runtime["structure"]["distinct_leaf_rows_observed"], "notes": ""},
        {"section": "runtime", "metric": "no_data_detected", "value": runtime["no_data_detected"], "notes": ""},
        {"section": "runtime", "metric": "message_text", "value": runtime["message_text"], "notes": ""},
        {"section": "support", "metric": "accepted_api_checked", "value": curated["accepted_api_support"]["checked"], "notes": ""},
        {"section": "support", "metric": "accepted_api_note", "value": curated["accepted_api_support"]["note"], "notes": ""},
        {"section": "boundary", "metric": "report_equivalence_status", "value": sc["report_equivalence_status"], "notes": ""},
        {"section": "boundary", "metric": "safe_claim_boundary", "value": sc["safe_claim_boundary"], "notes": ""},
    ]


def build_normalized_rows(curated: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "report_code": curated["report_code"],
            "report_name": curated["report_name"],
            "date_start": curated["date_start"],
            "date_end": curated["date_end"],
            "row_source": "runtime_html",
            "row_index": None,
            "service_name": None,
            "invoice_id": None,
            "patient_name": None,
            "amount": None,
            "runtime_support_status": "no_data_runtime_slice",
            "notes": "runtime report opened but rendered no data and no business table on this date",
        }
    ]


def build_warning_rows(curated: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "warning_code": item["code"],
            "severity": item["severity"],
            "layer": "rpt39",
            "message": item["message"],
        }
        for item in curated.get("warnings", [])
    ]


def build_notes_rows(date: str) -> list[dict[str, Any]]:
    return [
        {"field": "live_report_md", "value": (REPORTS_DIR / f"RPT_39_live_report_{date}.md").read_text()},
        {"field": "problem_log_md", "value": (REPORTS_DIR / f"RPT_39_problem_log_{date}.md").read_text()},
    ]


def build_meta(curated: dict[str, Any], output_path: Path) -> dict[str, Any]:
    return {
        "report_code": curated["report_code"],
        "report_name": curated["report_name"],
        "date_start": curated["date_start"],
        "date_end": curated["date_end"],
        "generated_at": curated["generated_at"],
        "run_status": curated["run_status"],
        "primary_truth": curated["source_classification"]["primary_truth"],
        "runtime_checked": True,
        "export_checked": False,
        "export_available": False,
        "runtime_row_count": curated["runtime"]["structure"]["distinct_leaf_rows_observed"],
        "export_row_count": None,
        "parity_status": "not_checked",
        "warnings": curated["warnings"],
        "safe_claim_boundary": curated["source_classification"]["safe_claim_boundary"],
        "normalized_workbook_path": str(output_path),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default="2026-02-18")
    parser.add_argument("--output")
    args = parser.parse_args()

    date = args.date
    curated = read_json(REPORTING_DIR / f"RPT_39_live_probe_{date}.json")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "RPT_39_summary"
    write_table(ws_summary, 1, build_summary_rows(curated), title="summary")

    ws_norm = wb.create_sheet("RPT_39_normalized")
    write_table(ws_norm, 1, build_normalized_rows(curated), title="normalized_rows")

    ws_warn = wb.create_sheet("RPT_39_warnings")
    end_row = write_table(ws_warn, 1, build_warning_rows(curated), title="warnings")
    for row in ws_warn.iter_rows(min_row=3, max_row=end_row):
        for cell in row:
            if cell.value is not None:
                cell.fill = WARN_FILL

    ws_notes = wb.create_sheet("RPT_39_notes")
    write_table(ws_notes, 1, build_notes_rows(date), title="notes")

    for ws in wb.worksheets:
        autofit(ws)

    output_path = Path(args.output) if args.output else EXCEL_DIR / f"RPT_39_{date}_normalized.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    load_workbook(output_path)

    meta_path = ARTIFACTS_DIR / f"RPT_39_{date}_meta.json"
    meta_path.write_text(json.dumps(build_meta(curated, output_path), ensure_ascii=False, indent=2) + "\n")
    print(output_path)
    print(meta_path)


if __name__ == "__main__":
    main()
