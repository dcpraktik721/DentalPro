#!/usr/bin/env python3
"""Build normalized workbook and meta sidecar for RPT_24 from saved UI/API artifacts."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path("/Users/macbook15/Downloads/MacAi/DentalPro")
REPORTING_DIR = REPO_ROOT / "reporting"
REPORTS_DIR = REPO_ROOT / "reports"
EXCEL_DIR = REPO_ROOT / "excel"
ARTIFACTS_DIR = REPO_ROOT / "artifacts"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EFEFEF")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
BOLD = Font(bold=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text())


def stringify(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width = len(str(cell.value))
            widths[cell.column] = min(max(widths.get(cell.column, 0), width), 80)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width + 2, 12)


def write_table(ws, start_row: int, rows: list[dict[str, Any]], title: str | None = None) -> int:
    current_row = start_row
    if title:
        ws.cell(current_row, 1, title).font = BOLD
        ws.cell(current_row, 1).fill = SECTION_FILL
        current_row += 1
    if not rows:
        ws.cell(current_row, 1, "no_rows")
        return current_row + 2
    headers: list[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
    for row_idx, row in enumerate(rows, start=current_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, stringify(row.get(header)))
    return current_row + len(rows) + 2


def build_meta(ui_artifact: dict[str, Any], api_artifact: dict[str, Any], workbook_path: Path) -> dict[str, Any]:
    derived = api_artifact["derived"]
    totals_cmp = derived["totals_comparison"]
    warnings = [
        {
            "code": "XRAY_PERFORMER_KEY_NOT_PROVEN",
            "severity": "high",
            "layer": "accepted_api_support",
            "message": "Xray performer row-level key is still not proven; performer mapping remains partial for that subfamily.",
        },
        {
            "code": "EXPORT_NOT_CHECKED",
            "severity": "medium",
            "layer": "export",
            "message": "No native export layer was reopened for this slice; runtime UI and accepted API composite are the checked layers.",
        },
    ]
    return {
        "report_code": "RPT_24",
        "report_name": "Выручка по отделениям",
        "date_start": api_artifact["date_start"],
        "date_end": api_artifact["date_end"],
        "generated_at": api_artifact["generated_at"],
        "run_status": "success_with_warnings" if totals_cmp["pass"] else "partial",
        "primary_truth": "runtime_html_validated_with_accepted_api_composite",
        "runtime_checked": True,
        "export_checked": False,
        "export_available": False,
        "runtime_row_count": derived["ui_report"]["row_count"] if derived["ui_report"] else None,
        "api_row_count": derived["totals"]["row_count"],
        "export_row_count": None,
        "parity_status": "totals_confirmed_with_known_xray_gap" if totals_cmp["pass"] else "mismatch_detected",
        "warnings": warnings,
        "safe_claim_boundary": "For 2026-02-18, RPT_24 is usable with limits: runtime UI and accepted API composite agree on row count and totals, but xray performer row-level linkage remains unresolved. The accepted API composite must not be promoted to fully canonical report truth for all subfamilies.",
        "ui_artifact_path": str(REPORTING_DIR / f"report-24-{api_artifact['date_start']}-{api_artifact['date_end']}.json"),
        "api_artifact_path": str(REPORTING_DIR / f"report-24-api-analysis-{api_artifact['date_start']}-{api_artifact['date_end']}.json"),
        "workbook_path": str(workbook_path),
    }


def build_summary_rows(ui_artifact: dict[str, Any], api_artifact: dict[str, Any], meta: dict[str, Any]) -> list[dict[str, Any]]:
    derived = api_artifact["derived"]
    totals = derived["totals"]
    totals_cmp = derived["totals_comparison"]
    return [
        {"section": "run", "metric": "report_code", "value": meta["report_code"], "notes": ""},
        {"section": "run", "metric": "report_name", "value": meta["report_name"], "notes": ""},
        {"section": "run", "metric": "date_start", "value": meta["date_start"], "notes": ""},
        {"section": "run", "metric": "date_end", "value": meta["date_end"], "notes": ""},
        {"section": "run", "metric": "run_status", "value": meta["run_status"], "notes": ""},
        {"section": "truth", "metric": "primary_truth", "value": meta["primary_truth"], "notes": ""},
        {"section": "runtime", "metric": "runtime_headers", "value": ui_artifact["flex_tables"][0]["headers"], "notes": ""},
        {"section": "runtime", "metric": "runtime_row_count", "value": derived["ui_report"]["row_count"], "notes": ""},
        {"section": "runtime", "metric": "runtime_total_paid", "value": derived["ui_report"]["totals"]["paid"], "notes": ""},
        {"section": "api", "metric": "final_rows", "value": totals["row_count"], "notes": ""},
        {"section": "api", "metric": "quantity", "value": totals["quantity"], "notes": ""},
        {"section": "api", "metric": "discount", "value": totals["discount"], "notes": ""},
        {"section": "api", "metric": "paid_ui", "value": totals["paid_ui"], "notes": ""},
        {"section": "api", "metric": "paid_raw", "value": totals["paid_raw"], "notes": ""},
        {"section": "api", "metric": "xray_rows", "value": len(derived["xray_rows"]), "notes": ""},
        {"section": "api", "metric": "supplemental_schedule_rows", "value": len(derived["supplemental_schedule_rows"]), "notes": ""},
        {"section": "verdict", "metric": "totals_pass", "value": totals_cmp["pass"], "notes": ""},
        {"section": "verdict", "metric": "xray_direct_key_found", "value": derived["xray_executor_evidence"]["direct_row_level_key_found"], "notes": ""},
        {"section": "boundary", "metric": "safe_claim_boundary", "value": meta["safe_claim_boundary"], "notes": ""},
    ]


def build_normalized_rows(api_artifact: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for idx, row in enumerate(api_artifact["derived"]["final_rows"], start=1):
        rows.append(
            {
                "report_code": "RPT_24",
                "report_name": "Выручка по отделениям",
                "date_start": api_artifact["date_start"],
                "date_end": api_artifact["date_end"],
                "row_source": row.get("source"),
                "row_index": idx,
                "pay_id": row.get("pay_id"),
                "object": row.get("object"),
                "doctor_id": row.get("doctor_id"),
                "executor": row.get("executor"),
                "client_id": row.get("client_id"),
                "patient_name": row.get("patient"),
                "phone": row.get("phone"),
                "date": row.get("date"),
                "service": row.get("service"),
                "quantity": row.get("quantity"),
                "discount_ui": row.get("discount_ui"),
                "paid_ui": row.get("paid_ui"),
                "payment_type": row.get("payment_type"),
                "insurance_with_discount": row.get("insurance_with_discount"),
                "corporate_zero_payment": row.get("corporate_zero_payment"),
                "xray_row": row.get("object") == "xray",
                "notes": "xray performer remains partial" if row.get("object") == "xray" else "",
            }
        )
    return rows


def build_warning_rows(meta: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "warning_code": item["code"],
            "severity": item["severity"],
            "layer": item["layer"],
            "message": item["message"],
        }
        for item in meta["warnings"]
    ]


def build_notes_rows(report_md: Path, problem_md: Path) -> list[dict[str, Any]]:
    return [
        {"field": "live_report_md", "value": report_md.read_text()},
        {"field": "problem_log_md", "value": problem_md.read_text()},
    ]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True)
    args = parser.parse_args()

    date = args.date
    ui_artifact = read_json(REPORTING_DIR / f"report-24-{date}-{date}.json")
    api_artifact = read_json(REPORTING_DIR / f"report-24-api-analysis-{date}-{date}.json")
    output_path = EXCEL_DIR / f"RPT_24_{date}_normalized.xlsx"
    meta_path = ARTIFACTS_DIR / f"RPT_24_{date}_meta.json"
    report_md = REPORTS_DIR / f"RPT_24_live_report_{date}.md"
    problem_md = REPORTS_DIR / f"RPT_24_problem_log_{date}.md"

    meta = build_meta(ui_artifact, api_artifact, output_path)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "RPT_24_summary"
    write_table(ws_summary, 1, build_summary_rows(ui_artifact, api_artifact, meta), title="summary")

    ws_norm = wb.create_sheet("RPT_24_normalized")
    write_table(ws_norm, 1, build_normalized_rows(api_artifact), title="normalized_rows")

    ws_warn = wb.create_sheet("RPT_24_warnings")
    end_row = write_table(ws_warn, 1, build_warning_rows(meta), title="warnings")
    for row in ws_warn.iter_rows(min_row=3, max_row=end_row):
        for cell in row:
            if cell.value is not None:
                cell.fill = WARN_FILL

    ws_notes = wb.create_sheet("RPT_24_notes")
    write_table(ws_notes, 1, build_notes_rows(report_md, problem_md), title="notes")

    for ws in wb.worksheets:
        autofit(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    meta_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    load_workbook(output_path)
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2) + "\n")
    print(output_path)
    print(meta_path)


if __name__ == "__main__":
    main()
