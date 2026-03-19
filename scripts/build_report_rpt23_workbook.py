#!/usr/bin/env python3
"""Build normalized workbook and meta sidecar for RPT_23 from saved repo artifacts."""

from __future__ import annotations

import argparse
import json
from datetime import datetime, UTC
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path("/Users/macbook15/Downloads/MacAi/DentalPro")
REPORTING_DIR = REPO_ROOT / "reporting"
REPORTS_DIR = REPO_ROOT / "reports"
EXCEL_DIR = REPO_ROOT / "excel"
ARTIFACTS_DIR = REPO_ROOT / "artifacts"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EFEFEF")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
BOLD = Font(bold=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text())


def stringify(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width = len(str(cell.value))
            widths[cell.column] = min(max(widths.get(cell.column, 0), width), 80)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width + 2, 12)


def write_table(ws, start_row: int, rows: list[dict[str, Any]], title: str | None = None) -> int:
    current_row = start_row
    if title:
        ws.cell(current_row, 1, title).font = BOLD
        ws.cell(current_row, 1).fill = SECTION_FILL
        current_row += 1
    if not rows:
        ws.cell(current_row, 1, "no_rows")
        return current_row + 2
    headers: list[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
    for row_idx, row in enumerate(rows, start=current_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, stringify(row.get(header)))
    return current_row + len(rows) + 2


def parse_money(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("₽", "").replace(" ", "").replace("\u00a0", "").replace(",", ".").strip()
    if not text:
        return None
    return float(text)


def parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    return int(str(value).replace(" ", "").strip())


def split_service_label(value: str) -> tuple[str | None, str]:
    if value.startswith("[") and "] " in value:
        prefix, label = value.split("] ", 1)
        return prefix + "]", label
    return None, value


def build_summary_rows(runtime: dict[str, Any], export: dict[str, Any], file_parse: dict[str, Any], verdict: dict[str, Any], meta: dict[str, Any]) -> list[dict[str, Any]]:
    footer = (runtime.get("runtime_footer_rows") or [{}])[0]
    return [
        {"section": "run", "metric": "report_code", "value": meta["report_code"], "notes": ""},
        {"section": "run", "metric": "report_name", "value": meta["report_name"], "notes": ""},
        {"section": "run", "metric": "date_start", "value": meta["date_start"], "notes": ""},
        {"section": "run", "metric": "date_end", "value": meta["date_end"], "notes": ""},
        {"section": "run", "metric": "run_status", "value": meta["run_status"], "notes": ""},
        {"section": "truth", "metric": "primary_truth", "value": meta["primary_truth"], "notes": ""},
        {"section": "truth", "metric": "safe_claim_boundary", "value": meta["safe_claim_boundary"], "notes": ""},
        {"section": "runtime", "metric": "runtime_row_count", "value": runtime["runtime_row_count"], "notes": ""},
        {"section": "runtime", "metric": "runtime_headers", "value": runtime["runtime_headers_exact"], "notes": ""},
        {"section": "runtime_footer", "metric": "avg_price", "value": parse_money(footer.get("Средняя цена")), "notes": "aggregate footer excluded from row parity"},
        {"section": "runtime_footer", "metric": "quantity", "value": parse_int(footer.get("Кол-во")), "notes": ""},
        {"section": "runtime_footer", "metric": "sum_done", "value": parse_money(footer.get("Сумма выполнено")), "notes": ""},
        {"section": "runtime_footer", "metric": "sum_discount", "value": parse_money(footer.get("Сумма скидки")), "notes": ""},
        {"section": "runtime_footer", "metric": "sum_paid_with_discount", "value": parse_money(footer.get("Сумма оплачено с учетом скидки")), "notes": ""},
        {"section": "runtime_footer", "metric": "cash", "value": parse_money(footer.get("Наличные")), "notes": ""},
        {"section": "runtime_footer", "metric": "noncash", "value": parse_money(footer.get("Безналичные")), "notes": ""},
        {"section": "runtime_footer", "metric": "total", "value": parse_money(footer.get("Общая")), "notes": ""},
        {"section": "export", "metric": "export_checked", "value": meta["export_checked"], "notes": ""},
        {"section": "export", "metric": "export_available", "value": meta["export_available"], "notes": ""},
        {"section": "export", "metric": "export_mode", "value": export["export_mode"], "notes": ""},
        {"section": "export", "metric": "file_received", "value": export["file_received"], "notes": ""},
        {"section": "file", "metric": "file_row_count", "value": file_parse["file_row_count"], "notes": ""},
        {"section": "file", "metric": "file_headers", "value": file_parse["file_headers_exact"], "notes": ""},
        {"section": "file", "metric": "file_schema_status", "value": file_parse["file_schema_status"], "notes": ""},
        {"section": "file", "metric": "file_row_boundary_status", "value": file_parse["file_row_boundary_status"], "notes": ""},
        {"section": "verdict", "metric": "family_class", "value": verdict["family_class"], "notes": ""},
        {"section": "verdict", "metric": "raw_parity_status", "value": verdict["raw_parity_status"], "notes": ""},
        {"section": "verdict", "metric": "normalized_business_parity_status", "value": verdict["normalized_business_parity_status"], "notes": ""},
        {"section": "verdict", "metric": "canonicalization_status", "value": verdict["canonicalization_status"], "notes": ""},
        {"section": "verdict", "metric": "safe_for_future_reporting", "value": verdict["safe_for_future_reporting"], "notes": ""},
    ]


def build_normalized_rows(runtime: dict[str, Any], file_parse: dict[str, Any]) -> list[dict[str, Any]]:
    file_rows = file_parse.get("rows") or []
    normalized = []
    for idx, runtime_row in enumerate(runtime["runtime_rows_structured"], start=1):
        file_row = file_rows[idx - 1] if idx - 1 < len(file_rows) else {}
        raw_service = runtime_row.get("Услуга") or ""
        service_code, service_label = split_service_label(raw_service)
        normalized.append(
            {
                "report_code": "RPT_23",
                "report_name": "Услуги за период",
                "date_start": runtime["target_date"],
                "date_end": runtime["target_date"],
                "row_source": "runtime_html",
                "row_index": idx,
                "runtime_service_raw": raw_service,
                "runtime_service_code_prefix": service_code,
                "runtime_service_label": service_label,
                "file_service_label": file_row.get("Услуга"),
                "service_first_column_exact_match": raw_service == file_row.get("Услуга"),
                "service_label_match_after_prefix_strip": service_label == file_row.get("Услуга"),
                "avg_price": parse_money(runtime_row.get("Средняя цена")),
                "quantity": parse_int(runtime_row.get("Кол-во")),
                "sum_done": parse_money(runtime_row.get("Сумма выполнено")),
                "sum_discount": parse_money(runtime_row.get("Сумма скидки")),
                "sum_paid_with_discount": parse_money(runtime_row.get("Сумма оплачено с учетом скидки")),
                "cash": parse_money(runtime_row.get("Наличные")),
                "noncash": parse_money(runtime_row.get("Безналичные")),
                "total": parse_money(runtime_row.get("Общая")),
                "file_avg_price": file_row.get("Средняя цена"),
                "file_quantity": file_row.get("Кол-во"),
                "file_sum_done": file_row.get("Сумма выполнено"),
                "file_sum_discount": file_row.get("Сумма скидки"),
                "file_sum_paid_with_discount": file_row.get("Сумма оплачено с учетом скидки"),
                "file_cash": file_row.get("Наличные"),
                "file_noncash": file_row.get("Безналичные"),
                "file_total": file_row.get("Общая"),
                "numeric_columns_match": all(
                    [
                        parse_money(runtime_row.get("Средняя цена")) == (float(file_row.get("Средняя цена")) if file_row.get("Средняя цена") is not None else None),
                        parse_int(runtime_row.get("Кол-во")) == file_row.get("Кол-во"),
                        parse_money(runtime_row.get("Сумма выполнено")) == (float(file_row.get("Сумма выполнено")) if file_row.get("Сумма выполнено") is not None else None),
                        parse_money(runtime_row.get("Сумма скидки")) == (float(file_row.get("Сумма скидки")) if file_row.get("Сумма скидки") is not None else None),
                        parse_money(runtime_row.get("Сумма оплачено с учетом скидки")) == (float(file_row.get("Сумма оплачено с учетом скидки")) if file_row.get("Сумма оплачено с учетом скидки") is not None else None),
                        parse_money(runtime_row.get("Наличные")) == (float(file_row.get("Наличные")) if file_row.get("Наличные") is not None else None),
                        parse_money(runtime_row.get("Безналичные")) == (float(file_row.get("Безналичные")) if file_row.get("Безналичные") is not None else None),
                        parse_money(runtime_row.get("Общая")) == (float(file_row.get("Общая")) if file_row.get("Общая") is not None else None),
                    ]
                ),
                "parity_status": "partial_due_to_first_column_policy",
                "notes": "runtime remains primary truth; first-column canonicalization still not closed",
            }
        )
    return normalized


def build_file_support_rows(file_parse: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for idx, row in enumerate(file_parse.get("rows", []), start=1):
        rows.append({"row_index": idx, **row})
    return rows


def build_warning_rows(verdict: dict[str, Any], export: dict[str, Any]) -> list[dict[str, Any]]:
    warnings = [
        {
            "warning_code": "RAW_PARITY_FAILED",
            "severity": "high",
            "layer": "runtime_vs_export_file",
            "message": "Raw parity failed; row-count equality is not sufficient for final parity on this slice.",
        },
        {
            "warning_code": "NORMALIZED_PARITY_PARTIAL",
            "severity": "high",
            "layer": "governed_verdict",
            "message": "Normalized business parity is only partial because first-column policy is unresolved.",
        },
        {
            "warning_code": "STILL_NOT_CANONICAL",
            "severity": "high",
            "layer": "canonicalization",
            "message": "Canonicalization remains still_not_canonical; do not collapse runtime labels into file labels silently.",
        },
        {
            "warning_code": "RUNTIME_FIRST_REQUIRED",
            "severity": "medium",
            "layer": "truth_boundary",
            "message": "Runtime HTML remains the primary truth layer; export/file is secondary audit support only.",
        },
    ]
    if not export.get("file_received"):
        warnings.append(
            {
                "warning_code": "EXPORT_FILE_NOT_RECEIVED",
                "severity": "high",
                "layer": "export",
                "message": "Native export was not received for this run.",
            }
        )
    return warnings


def build_notes_rows(report_md: Path, problem_md: Path, technical_md: Path | None = None) -> list[dict[str, Any]]:
    rows = [
        {"field": "live_report_md", "value": report_md.read_text()},
        {"field": "problem_log_md", "value": problem_md.read_text()},
    ]
    if technical_md and technical_md.exists():
        rows.append({"field": "external_technical_disclosure_md", "value": technical_md.read_text()})
    return rows


def build_meta(runtime: dict[str, Any], export: dict[str, Any], file_parse: dict[str, Any], verdict: dict[str, Any], workbook_path: Path) -> dict[str, Any]:
    warnings = build_warning_rows(verdict, export)
    return {
        "report_code": "RPT_23",
        "report_name": "Услуги за период",
        "date_start": runtime["target_date"],
        "date_end": runtime["target_date"],
        "generated_at": datetime.now(UTC).isoformat(),
        "run_status": "success_with_warnings",
        "primary_truth": "runtime_html",
        "runtime_checked": True,
        "export_checked": True,
        "export_available": export["export_available"] == "yes",
        "runtime_row_count": runtime["runtime_row_count"],
        "export_row_count": file_parse["file_row_count"],
        "parity_status": verdict["normalized_business_parity_status"],
        "warnings": warnings,
        "safe_claim_boundary": "For 2026-02-18, RPT_23 is runtime-first. Native export/file is a secondary audit layer. Raw parity failed, normalized business parity is partial, and first-column canonicalization remains unresolved. Do not claim full parity or canonicalization.",
        "export_file_path": str(ARTIFACTS_DIR / "RPT_23_2026-02-18_native_export.xlsx"),
        "runtime_artifact_path": str(REPORTING_DIR / "RPT_23_runtime_live_2026-02-18.json"),
        "file_parse_artifact_path": str(REPORTING_DIR / "RPT_23_file_parse_2026-02-18.json"),
        "governed_verdict_path": str(REPORTING_DIR / "RPT_23_final_governed_verdict_2026-02-18.json"),
        "workbook_path": str(workbook_path),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True)
    args = parser.parse_args()

    date = args.date
    runtime = read_json(REPORTING_DIR / f"RPT_23_runtime_live_{date}.json")
    export = read_json(REPORTING_DIR / f"RPT_23_export_live_{date}.json")
    file_parse = read_json(REPORTING_DIR / f"RPT_23_file_parse_{date}.json")
    verdict = read_json(REPORTING_DIR / f"RPT_23_final_governed_verdict_{date}.json")

    output_path = EXCEL_DIR / f"RPT_23_{date}_normalized.xlsx"
    meta_path = ARTIFACTS_DIR / f"RPT_23_{date}_meta.json"
    report_md = REPORTS_DIR / f"RPT_23_live_report_{date}.md"
    problem_md = REPORTS_DIR / f"RPT_23_problem_log_{date}.md"
    technical_md = Path("/Users/macbook15/Downloads/YDIREKT_code/mac-ai-os/dentalpro-playwright/docs") / f"RPT_23_technical_disclosure_{date}.md"

    meta = build_meta(runtime, export, file_parse, verdict, output_path)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "RPT_23_summary"
    write_table(ws_summary, 1, build_summary_rows(runtime, export, file_parse, verdict, meta), title="summary")

    ws_norm = wb.create_sheet("RPT_23_normalized")
    write_table(ws_norm, 1, build_normalized_rows(runtime, file_parse), title="normalized_rows")

    ws_file = wb.create_sheet("RPT_23_file_support")
    write_table(ws_file, 1, build_file_support_rows(file_parse), title="file_support_rows")

    ws_warn = wb.create_sheet("RPT_23_warnings")
    end_row = write_table(ws_warn, 1, build_warning_rows(verdict, export), title="warnings")
    for row in ws_warn.iter_rows(min_row=3, max_row=end_row):
        for cell in row:
            if cell.value is not None:
                cell.fill = WARN_FILL

    ws_notes = wb.create_sheet("RPT_23_notes")
    write_table(ws_notes, 1, build_notes_rows(report_md, problem_md, technical_md), title="notes")

    for ws in wb.worksheets:
        autofit(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    meta_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    load_workbook(output_path)
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2) + "\n")
    print(output_path)
    print(meta_path)


if __name__ == "__main__":
    main()
