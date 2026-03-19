#!/usr/bin/env python3
"""Build a normalized Excel workbook for RPT_9 from saved repo artifacts."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path("/Users/macbook15/Downloads/MacAi/DentalPro")
REPORTING_DIR = REPO_ROOT / "reporting"
REPORTS_DIR = REPO_ROOT / "reports"
EXCEL_DIR = REPO_ROOT / "excel"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EFEFEF")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
BOLD = Font(bold=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text())


def stringify(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width = len(str(cell.value))
            widths[cell.column] = min(max(widths.get(cell.column, 0), width), 80)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width + 2, 12)


def write_table(ws, start_row: int, rows: list[dict[str, Any]], title: str | None = None) -> int:
    current_row = start_row
    if title:
        ws.cell(current_row, 1, title).font = BOLD
        ws.cell(current_row, 1).fill = SECTION_FILL
        current_row += 1

    if not rows:
        ws.cell(current_row, 1, "no_rows")
        return current_row + 2

    headers: list[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    for row_idx, row in enumerate(rows, start=current_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, stringify(row.get(header)))

    return current_row + len(rows) + 2


def build_summary_rows(curated: dict[str, Any], raw_api: dict[str, Any]) -> list[dict[str, Any]]:
    runtime = curated["runtime"]
    api_support = curated["accepted_api_support"]
    auth = curated["auth"]
    sc = curated["source_classification"]
    group_row = runtime.get("observed_group_row") or {}
    target_invoice = api_support.get("targeted_invoice_detail") or {}
    target_pay = api_support.get("targeted_z_pays_match") or {}
    return [
        {"section": "run", "metric": "report_code", "value": curated["report_code"], "notes": ""},
        {"section": "run", "metric": "report_name", "value": curated["report_name"], "notes": ""},
        {"section": "run", "metric": "date_start", "value": curated["date_start"], "notes": ""},
        {"section": "run", "metric": "date_end", "value": curated["date_end"], "notes": ""},
        {"section": "run", "metric": "generated_at", "value": curated["generated_at"], "notes": ""},
        {"section": "run", "metric": "run_status", "value": curated["run_status"], "notes": ""},
        {"section": "auth", "metric": "authenticated", "value": auth["authenticated"], "notes": ""},
        {"section": "auth", "metric": "final_url", "value": auth["final_url"], "notes": ""},
        {"section": "auth", "metric": "success_reason", "value": auth["success_reason"], "notes": ""},
        {"section": "runtime", "metric": "primary_truth", "value": sc["primary_truth"], "notes": ""},
        {"section": "runtime", "metric": "table_rendering", "value": runtime["structure"]["table_rendering"], "notes": ""},
        {"section": "runtime", "metric": "distinct_leaf_rows_observed", "value": runtime["structure"]["distinct_leaf_rows_observed"], "notes": ""},
        {"section": "runtime", "metric": "observed_group_levels", "value": " -> ".join(runtime["structure"]["group_levels"]), "notes": ""},
        {"section": "runtime", "metric": "branch_group", "value": group_row.get("branch"), "notes": ""},
        {"section": "runtime", "metric": "status_group", "value": group_row.get("status_group"), "notes": ""},
        {"section": "runtime", "metric": "creator_group", "value": group_row.get("creator_group"), "notes": ""},
        {"section": "runtime", "metric": "patient_group", "value": group_row.get("patient_group"), "notes": ""},
        {"section": "runtime", "metric": "leaf_invoice_id", "value": (runtime.get("observed_leaf_row") or {}).get("cashbox_form_id"), "notes": ""},
        {"section": "runtime", "metric": "leaf_patient_id", "value": (runtime.get("observed_leaf_row") or {}).get("patient_id"), "notes": ""},
        {"section": "runtime", "metric": "leaf_amount_due", "value": (runtime.get("observed_leaf_row") or {}).get("amount_due"), "notes": ""},
        {"section": "api", "metric": "z_pays_date_row_count", "value": api_support["z_pays_date_row_count"], "notes": "broader than report output"},
        {"section": "api", "metric": "z_pays_status_counts", "value": api_support["z_pays_status_counts"], "notes": ""},
        {"section": "api", "metric": "target_invoice_status_name", "value": target_invoice.get("status_name"), "notes": ""},
        {"section": "api", "metric": "target_z_pays_status_text", "value": target_pay.get("status_text"), "notes": ""},
        {"section": "api", "metric": "missing_fields", "value": api_support["missing_fields_in_accepted_api"], "notes": "not proven via accepted APIs"},
        {"section": "boundary", "metric": "report_equivalence_status", "value": sc["report_equivalence_status"], "notes": ""},
        {"section": "boundary", "metric": "safe_claim_boundary", "value": sc["safe_claim_boundary"], "notes": ""},
        {"section": "api_raw", "metric": "methods_checked", "value": raw_api.get("methods_checked", []), "notes": ""},
    ]


def build_normalized_rows(curated: dict[str, Any]) -> list[dict[str, Any]]:
    row = curated["runtime"].get("observed_leaf_row")
    api = curated.get("accepted_api_support", {})
    if not row:
        return [
            {
                "report_code": curated["report_code"],
                "report_name": curated["report_name"],
                "date_start": curated["date_start"],
                "date_end": curated["date_end"],
                "row_source": "runtime_html",
                "row_index": None,
                "patient_name": None,
                "patient_id": None,
                "invoice_id": None,
                "invoice_name": None,
                "branch": None,
                "creator_name": None,
                "created_at": None,
                "changed_at": None,
                "reason": None,
                "amount_due": None,
                "status_group": None,
                "invoice_detail_status_name": None,
                "z_pays_status_text": None,
                "runtime_api_identity_match": None,
                "updater_history_semantics_status": "not_proven",
                "notes": "runtime report rendered with no leaf rows for this date",
            }
        ]
    return [
        {
            "report_code": curated["report_code"],
            "report_name": curated["report_name"],
            "date_start": curated["date_start"],
            "date_end": curated["date_end"],
            "row_source": "runtime_html",
            "row_index": 1,
            "patient_name": row["patient_name"],
            "patient_id": row["patient_id"],
            "invoice_id": row["cashbox_form_id"],
            "invoice_name": row["invoice_name"],
            "branch": row["branch"],
            "creator_name": row["creator_name"],
            "created_at": row["created_at"],
            "changed_at": row["changed_at"],
            "reason": row["reason"],
            "amount_due": row["amount_due"],
            "status_group": row["status_group"],
            "invoice_detail_status_name": api.get("targeted_invoice_detail", {}).get("status_name"),
            "z_pays_status_text": api.get("targeted_z_pays_match", {}).get("status_text"),
            "runtime_api_identity_match": True,
            "updater_history_semantics_status": "not_proven",
            "notes": "single observed leaf row for this date",
        }
    ]


def build_group_rows(curated: dict[str, Any]) -> list[dict[str, Any]]:
    group = curated["runtime"].get("observed_group_row")
    structure = curated["runtime"]["structure"]
    if not group:
        return [
            {
                "group_level_index": None,
                "group_level_name": "no_groups",
                "group_value": None,
                "group_amount": None,
                "notes": "runtime report contained no grouped business rows on this date",
            }
        ]
    rows = []
    for idx, level in enumerate(structure["group_levels"], start=1):
        key = {
            "Филиал": "branch",
            "Статус квитанции": "status_group",
            "Создатель квитанции": "creator_group",
            "Пациент": "patient_group",
        }.get(level)
        rows.append(
            {
                "group_level_index": idx,
                "group_level_name": level,
                "group_value": group.get(key, ""),
                "group_amount": group["group_amount"],
                "notes": "observed in nested runtime accordion structure",
            }
        )
    return rows


def build_api_rows(curated: dict[str, Any], raw_api: dict[str, Any]) -> list[dict[str, Any]]:
    target_invoice = curated["accepted_api_support"].get("targeted_invoice_detail")
    target_pay = curated["accepted_api_support"].get("targeted_z_pays_match")
    rows = [
        {
            "layer": "z_pays_summary",
            "field": "row_count",
            "value": curated["accepted_api_support"]["z_pays_date_row_count"],
            "notes": "date-only API output",
        },
        {
            "layer": "z_pays_summary",
            "field": "status_counts",
            "value": curated["accepted_api_support"]["z_pays_status_counts"],
            "notes": "",
        },
    ]
    if target_invoice:
        for key, value in target_invoice.items():
            rows.append({"layer": "invoice_detail_targeted", "field": key, "value": value, "notes": "targeted invoice"})
    if target_pay:
        for key, value in target_pay.items():
            rows.append({"layer": "z_pays_targeted", "field": key, "value": value, "notes": "targeted invoice"})
    for missing in curated["accepted_api_support"]["missing_fields_in_accepted_api"]:
        rows.append({"layer": "accepted_api_gap", "field": missing, "value": "missing", "notes": "not exposed as direct accepted read field"})
    if raw_api["invoice_detail"].get("sample_lookups"):
        for sample in raw_api["invoice_detail"]["sample_lookups"][:3]:
            rows.append(
                {
                    "layer": "invoice_detail_sample",
                    "field": sample.get("id"),
                    "value": sample.get("sample", {}),
                    "notes": "sample lookup from legacy analysis artifact",
                }
            )
    return rows


def build_warning_rows(curated: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for item in curated.get("warnings", []):
        rows.append(
            {
                "warning_code": item["code"],
                "severity": item["severity"],
                "layer": "rpt9",
                "message": item["message"],
            }
        )
    return rows


def build_notes_rows(date: str) -> list[dict[str, Any]]:
    return [
        {
            "field": "live_report_md",
            "value": (REPORTS_DIR / f"RPT_9_live_report_{date}.md").read_text(),
        },
        {
            "field": "problem_log_md",
            "value": (REPORTS_DIR / f"RPT_9_problem_log_{date}.md").read_text(),
        },
    ]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default="2026-03-18")
    parser.add_argument("--output")
    args = parser.parse_args()

    date = args.date
    curated = read_json(REPORTING_DIR / f"RPT_9_live_probe_{date}.json")
    raw_api = read_json(REPORTING_DIR / f"report-9-api-analysis-{date}-{date}.json")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "RPT_9_summary"
    write_table(ws_summary, 1, build_summary_rows(curated, raw_api), title="summary")

    ws_norm = wb.create_sheet("RPT_9_normalized")
    write_table(ws_norm, 1, build_normalized_rows(curated), title="normalized_rows")

    ws_groups = wb.create_sheet("RPT_9_groups")
    write_table(ws_groups, 1, build_group_rows(curated), title="observed_group_structure")

    ws_api = wb.create_sheet("RPT_9_api_support")
    write_table(ws_api, 1, build_api_rows(curated, raw_api), title="api_support")

    ws_warn = wb.create_sheet("RPT_9_warnings")
    end_row = write_table(ws_warn, 1, build_warning_rows(curated), title="warnings")
    for row in ws_warn.iter_rows(min_row=3, max_row=end_row):
        for cell in row:
            if cell.value is not None:
                cell.fill = WARN_FILL

    ws_notes = wb.create_sheet("RPT_9_notes")
    write_table(ws_notes, 1, build_notes_rows(date), title="notes")

    for ws in wb.worksheets:
        autofit(ws)

    output_path = Path(args.output) if args.output else EXCEL_DIR / f"RPT_9_{date}_normalized.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    load_workbook(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
