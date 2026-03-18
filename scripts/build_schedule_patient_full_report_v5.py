import json
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import re

TARGET_DATE = "2026-03-10"
TARGET_DATE_RU = "10.03.2026"
ART = Path("/Users/macbook15/Downloads/MacAi/artifacts")

VISITS = json.loads((ART / f"schedule_{TARGET_DATE}_visits.json").read_text())
STATUS_V2 = json.loads((ART / f"schedule_{TARGET_DATE}_status_dictionary_v2.json").read_text())
PATIENT_V5 = json.loads((ART / f"schedule_{TARGET_DATE}_patient_identity_probe_v5.json").read_text())
TRANSPORT_V2 = json.loads((ART / f"schedule_{TARGET_DATE}_transport_detail_candidates_v2.json").read_text())
PARSED_EXPORTS = json.loads((ART / "date_constrained_xlsx_parsed_2026_03_10.json").read_text())

OUT_XLSX = ART / f"schedule_{TARGET_DATE}_patient_full_report_v5.xlsx"
OUT_JSON = ART / f"schedule_{TARGET_DATE}_patient_full_report_v5.json"
OUT_CASH = ART / f"schedule_{TARGET_DATE}_cash_enrichment_v5.json"
OUT_MD = ART / f"schedule_{TARGET_DATE}_patient_full_report_v5.md"


def normalize_name(value):
    return re.sub(r"\s+", " ", str(value or "")).strip().upper()


def normalize_phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    return digits


def surname(value):
    return normalize_name(value).split(" ")[0] if value else ""


status_by_pattern = {item["raw_icon_class_pattern"]: item for item in STATUS_V2}
patient_by_visit = {item["visit_id"]: item for item in PATIENT_V5}

rpt24_meta = next(item for item in PARSED_EXPORTS if item["report_code"] == "RPT_24")
wb24 = load_workbook(rpt24_meta["workbook_path"], data_only=True, read_only=True)
ws24 = wb24[rpt24_meta["sheet_name"]]
headers = [cell.value for cell in ws24[rpt24_meta["header_row_index"]]]
rows24 = []
for row_idx in range(rpt24_meta["data_row_start"], rpt24_meta["data_row_end"] + 1):
    values = [ws24.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
    rows24.append({headers[i]: values[i] for i in range(len(headers))})

from collections import defaultdict
rpt24_index = defaultdict(list)
for row in rows24:
    key = (
        str(row.get("Дата", "")).strip(),
        normalize_name(row.get("Пациент", "")),
        normalize_phone(row.get("Номер телефона", "")),
    )
    rpt24_index[key].append(row)

main_rows = []
cash_rows = []
for visit in VISITS:
    pattern = visit.get("status_code", "")
    status = status_by_pattern.get(pattern, {})
    fields = status.get("normalized_fields", {})
    proofs = status.get("proof_levels", {})
    identity = patient_by_visit.get(visit["visit_id"], {})

    if visit["patient_name"] == "Резерв для пациента":
        task_family = "reserve_slot"
    elif visit["patient_name"] == "Обед":
        task_family = "service_block"
    elif (visit.get("status_label") or "") == "Онлайн запись":
        task_family = "online_booking"
    elif (visit.get("status_label") or "") == "Неявка":
        task_family = "no_show"
    else:
        task_family = "patient_visit"

    patient_full = identity.get("patient_name_full", "")
    patient_phone = identity.get("patient_phone_primary", "")
    patient_id = identity.get("patient_id", "")
    cash_key = (TARGET_DATE_RU, normalize_name(patient_full), normalize_phone(patient_phone))
    matched_cash = rpt24_index.get(cash_key, []) if patient_full and patient_phone else []
    matched_doctors = sorted({str(r.get("Исполнитель", "")).strip() for r in matched_cash if r.get("Исполнитель")})
    schedule_doctor_surname = surname(visit.get("doctor_name", ""))
    matched_doctor_surnames = sorted({surname(name) for name in matched_doctors if name})
    if not matched_cash:
        join_status = "not_proven"
        linkage_quality = "no_match"
    elif len(matched_doctors) == 1 and schedule_doctor_surname in matched_doctor_surnames:
        join_status = "conditional_join"
        linkage_quality = "patient_day_doctor_aligned"
    elif schedule_doctor_surname in matched_doctor_surnames:
        join_status = "conditional_join"
        linkage_quality = "patient_day_multi_doctor_ambiguous"
    else:
        join_status = "heuristic_join"
        linkage_quality = "patient_day_name_phone_only"

    if task_family in ("reserve_slot", "service_block"):
        row_readiness = "non_patient_service_slot"
    elif patient_id and patient_full and patient_phone:
        row_readiness = "identity_ready_cash_separate"
    elif patient_full or patient_phone:
        row_readiness = "identity_partial"
    else:
        row_readiness = "runtime_only"

    main_rows.append({
        "visit_id": visit["visit_id"],
        "patient_id": patient_id,
        "patient_id_proof": identity.get("patient_id_proof", "not_proven"),
        "patient_card_url": identity.get("patient_card_url", ""),
        "patient_card_url_proof": identity.get("patient_card_url_proof", "not_proven"),
        "patient_name_full": patient_full,
        "patient_name_full_proof": identity.get("patient_name_full_proof", "not_proven"),
        "patient_name_short": visit.get("patient_name", ""),
        "patient_phone_primary": patient_phone,
        "patient_phone_secondary": identity.get("patient_phone_secondary", ""),
        "patient_phone_raw": identity.get("patient_phone_raw", ""),
        "patient_contact_proof": identity.get("patient_contact_proof", "not_proven"),
        "patient_contact_source": identity.get("patient_contact_source", ""),
        "doctor_name": visit.get("doctor_name", ""),
        "cabinet_name": visit.get("cabinet_name", ""),
        "date": visit.get("date", ""),
        "start_time": visit.get("start_time", ""),
        "end_time": visit.get("end_time", ""),
        "visit_type": visit.get("visit_type", ""),
        "task_family": task_family,
        "visit_presence_state": fields.get("visit_presence_state", ""),
        "payment_state": fields.get("payment_state", ""),
        "cash_amount_total": "",
        "ticket_state": fields.get("ticket_state", ""),
        "booking_origin": fields.get("booking_origin", ""),
        "popup_linkage_state": identity.get("popup_linkage_state", "not_reprobed"),
        "row_readiness": row_readiness,
        "notes": "; ".join(identity.get("notes", [])),
        "patient_identity_source_layer": identity.get("patient_identity_source_layer", ""),
        "patient_identity_proof_level": (
            "proven" if patient_id and patient_full else
            "partial" if patient_full or patient_phone else
            "not_proven"
        ),
        "patient_identity_linkage_method": identity.get("patient_identity_linkage_method", ""),
        "patient_contact_source_layer": "schedule_popup" if patient_phone else "",
        "patient_contact_proof_level": identity.get("patient_contact_proof", "not_proven"),
        "patient_contact_linkage_method": identity.get("patient_identity_linkage_method", "") if patient_phone else "",
        "payment_source_layer": "",
        "payment_proof_level": "not_proven_direct_visit_link",
        "payment_linkage_method": "",
        "popup_phone_detected": identity.get("popup_phone_detected", "no"),
        "popup_has_patient_info_button": identity.get("popup_has_patient_info_button", "no"),
        "popup_has_ambcard_button": identity.get("popup_has_ambcard_button", "no"),
        "popup_has_treatment_plan_button": identity.get("popup_has_treatment_plan_button", "no"),
        "popup_has_questionnaire_button": identity.get("popup_has_questionnaire_button", "no"),
        "popup_has_visit_result_action": identity.get("popup_has_visit_result_action", "no"),
        "popup_actions_raw": identity.get("popup_actions_raw", ""),
    })

    cash_rows.append({
        "visit_id": visit["visit_id"],
        "patient_id": patient_id,
        "patient_name_full": patient_full,
        "patient_phone_primary": patient_phone,
        "doctor_name_schedule": visit.get("doctor_name", ""),
        "date": visit.get("date", ""),
        "cash_amount_total_candidate": sum(float(r.get("Сумма оплачено с учетом скидки") or 0) for r in matched_cash) if matched_cash else "",
        "discount_total_candidate": sum(float(r.get("Сумма скидки") or 0) for r in matched_cash) if matched_cash else "",
        "payment_event_count": len(matched_cash),
        "matched_rpt24_doctors": "; ".join(matched_doctors),
        "matched_rpt24_row_count": len(matched_cash),
        "payment_source": "RPT_24 file-first export 2026-03-10" if matched_cash else "",
        "payment_linkage_method": "conditional patient-day match on exact full_name + primary_phone + date" if matched_cash else "",
        "cash_amount_proof": join_status,
        "source_layer": "file_first_export:RPT_24" if matched_cash else "",
        "proof_level": join_status,
        "join_status": join_status,
        "linkage_quality": linkage_quality,
        "linkage_method": "patient_day_name_phone_date" if matched_cash else "",
        "notes": "patient-day financial signal only; not safe as direct visit->cash fact" if matched_cash else "no exact patient-day RPT_24 match from proven identity fields",
    })

qc = {
    "rows_total": len(main_rows),
    "rows_without_full_name": sum(1 for r in main_rows if not r["patient_name_full"]),
    "rows_without_phone": sum(1 for r in main_rows if not r["patient_phone_primary"]),
    "rows_without_patient_id": sum(1 for r in main_rows if not r["patient_id"] and r["patient_id_proof"] == "not_proven"),
    "rows_patient_id_not_applicable": sum(1 for r in main_rows if r["patient_id_proof"] == "not_applicable_service_row"),
    "rows_without_cash_amount_total_main_sheet": len(main_rows),
    "rows_with_not_proven_or_not_reprobed_enrichment": sum(
        1 for r in main_rows if r["patient_identity_proof_level"] != "proven" or r["popup_linkage_state"] != "proven_popup_linkage"
    ),
    "cash_enrichment_rows_with_conditional_match": sum(1 for r in cash_rows if r["cash_amount_proof"] == "conditional_join"),
    "cash_enrichment_rows_with_heuristic_match": sum(1 for r in cash_rows if r["cash_amount_proof"] == "heuristic_join"),
    "cash_enrichment_rows_without_match": sum(1 for r in cash_rows if r["cash_amount_proof"] == "not_proven"),
}

payload = {"patient_schedule_full": main_rows, "cash_enrichment": cash_rows, "qc": qc}
OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
OUT_CASH.write_text(json.dumps(cash_rows, ensure_ascii=False, indent=2))

wb = Workbook()
ws = wb.active
ws.title = "summary_ru"
fill = PatternFill("solid", fgColor="1F4E78")
font = Font(color="FFFFFF", bold=True)
def style_header(sheet):
    for c in sheet[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in [
    ["Показатель", "Значение"],
    ["Дата расписания", TARGET_DATE],
    ["Всего записей", len(main_rows)],
    ["Proven patient_id", sum(1 for r in main_rows if r["patient_id"])],
    ["patient_id not applicable", qc["rows_patient_id_not_applicable"]],
    ["Conditional cash matches", qc["cash_enrichment_rows_with_conditional_match"]],
    ["Heuristic cash matches", qc["cash_enrichment_rows_with_heuristic_match"]],
]:
    ws.append(row)
style_header(ws)

for title, rows, headers in [
    ("patient_schedule_full", main_rows, list(main_rows[0].keys())),
    ("расписание_пациентов", main_rows, ["visit_id","patient_id","patient_name_full","patient_name_short","patient_phone_primary","doctor_name","cabinet_name","date","start_time","end_time","visit_type","task_family","visit_presence_state","payment_state","cash_amount_total","ticket_state","booking_origin","popup_linkage_state","row_readiness","notes"]),
    ("cash_enrichment", cash_rows, list(cash_rows[0].keys())),
]:
    sh = wb.create_sheet(title)
    sh.append(headers)
    for r in rows:
        if title == "расписание_пациентов":
            sh.append([r.get(h,"") for h in headers])
        else:
            sh.append([r.get(h,"") for h in headers])
    style_header(sh)

qc_sheet = wb.create_sheet("quality_control")
qc_sheet.append(["metric","value"])
for k,v in qc.items():
    qc_sheet.append([k,v])
style_header(qc_sheet)

for sh in wb.worksheets:
    for col in sh.columns:
        max_len = max(min(len(str(c.value or "")), 80) for c in col)
        sh.column_dimensions[get_column_letter(col[0].column)].width = max(14, max_len+2)
    sh.freeze_panes = "A2"

wb.save(OUT_XLSX)
OUT_MD.write_text(
    f"# Schedule {TARGET_DATE} Patient Full Report V5\\n\\n"
    f"- proven_patient_id: {sum(1 for r in main_rows if r['patient_id'])}\\n"
    f"- patient_id_not_applicable: {qc['rows_patient_id_not_applicable']}\\n"
    f"- unresolved_patient_identity_rows: {qc['rows_without_patient_id']}\\n"
    f"- conditional_cash_matches: {qc['cash_enrichment_rows_with_conditional_match']}\\n"
)
