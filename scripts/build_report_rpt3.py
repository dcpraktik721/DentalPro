#!/usr/bin/env python3
"""Production-safe builder for DentalPRO RPT_3 / Задолженности и авансы."""

from __future__ import annotations

import argparse
import base64
import copy
import json
import math
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode, urljoin

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
SPEC_PATH = REPO_ROOT / "registry" / "report_specs" / "rpt_3.json"
ARTIFACTS_DIR = REPO_ROOT / "artifacts"
EXCEL_DIR = REPO_ROOT / "excel"
RUNTIME_DIR = REPO_ROOT / "runtime"
AUTH_REFRESH_SCRIPT = RUNTIME_DIR / "auth_refresh.js"
AUTH_REFRESH_RESULT = RUNTIME_DIR / "auth_refresh_result.json"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EFEFEF")
WARN_FILL = PatternFill(fill_type="solid", fgColor="FFF3CD")
BOLD = Font(bold=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def looks_like_date(value: str) -> bool:
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", value))


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def normalize_date_for_filename(start: str, end: str) -> str:
    return start if start == end else f"{start}_{end}"


def normalize_phone(value: str | None) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D+", "", str(value))
    return digits or None


def parse_money(value: str | float | int | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace("\xa0", " ").replace("₽", "").replace(" ", "").replace(",", ".")
    if cleaned in {"", "-", "—"}:
        return None
    return float(cleaned)


def parse_ui_date(value: str | None) -> date | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return datetime.strptime(text, "%d.%m.%Y").date()


def stringify_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def sheet_autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width = len(str(cell.value))
            widths[cell.column] = min(max(widths.get(cell.column, 0), width), 80)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width + 2, 12)


def write_table(ws, start_row: int, rows: list[dict[str, Any]], title: str | None = None) -> int:
    row_idx = start_row
    if title:
        ws.cell(row_idx, 1, title).font = BOLD
        ws.cell(row_idx, 1).fill = SECTION_FILL
        row_idx += 1

    if not rows:
        ws.cell(row_idx, 1, "no_rows")
        return row_idx + 2

    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row_idx, col_idx, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    for row in rows:
        row_idx += 1
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, stringify_cell(row.get(header)))
    return row_idx + 2


def bool_close(lhs: float | None, rhs: float | None) -> bool:
    if lhs is None or rhs is None:
        return False
    return math.isclose(lhs, rhs, abs_tol=0.01)


def make_warning(code: str, severity: str, layer: str, message: str) -> dict[str, str]:
    return {
        "warning_code": code,
        "severity": severity,
        "layer": layer,
        "message": message,
    }


def relative_repo_path(path: Path) -> str:
    return str(path.relative_to(REPO_ROOT))


@dataclass
class ParsedInput:
    alias: str
    date_start: str
    date_end: str


def parse_cli(spec: dict[str, Any]) -> ParsedInput:
    parser = argparse.ArgumentParser(description="Build normalized Excel for DentalPRO RPT_3")
    parser.add_argument("positional", nargs="*")
    parser.add_argument("--date")
    parser.add_argument("--start")
    parser.add_argument("--end")
    parser.add_argument("--no-export", action="store_true")
    parser.add_argument("--skip-auth-refresh", action="store_true")
    parser.add_argument("--output-xlsx")
    parser.add_argument("--output-meta")
    args = parser.parse_args()

    aliases = set(spec["aliases"])
    script_stem = Path(sys.argv[0]).stem
    positional = list(args.positional)

    if script_stem in aliases:
        positional = [script_stem, *positional]

    alias = spec["report_code"]
    date_start: str | None = None
    date_end: str | None = None

    if args.date:
        date_start = args.date
        date_end = args.date
    elif args.start and args.end:
        date_start = args.start
        date_end = args.end
    else:
        if len(positional) == 1 and looks_like_date(positional[0]):
            date_start = positional[0]
            date_end = positional[0]
        elif len(positional) == 2:
            if positional[0] in aliases and looks_like_date(positional[1]):
                alias = positional[0]
                date_start = positional[1]
                date_end = positional[1]
            elif looks_like_date(positional[0]) and looks_like_date(positional[1]):
                date_start = positional[0]
                date_end = positional[1]
        elif len(positional) == 3 and positional[0] in aliases and looks_like_date(positional[1]) and looks_like_date(positional[2]):
            alias = positional[0]
            date_start = positional[1]
            date_end = positional[2]

    if alias not in aliases:
        raise SystemExit(f"Unsupported alias: {alias}")
    if not date_start or not date_end:
        raise SystemExit(
            "Usage: build_report_rpt3.py RPT_3 YYYY-MM-DD | build_report_rpt3.py RPT_3 YYYY-MM-DD YYYY-MM-DD | --date / --start --end"
        )
    if not looks_like_date(date_start) or not looks_like_date(date_end):
        raise SystemExit("Dates must be in YYYY-MM-DD format")

    return ParsedInput(alias=alias, date_start=date_start, date_end=date_end)


def get_output_paths(spec: dict[str, Any], date_start: str, date_end: str, output_xlsx: str | None, output_meta: str | None) -> tuple[Path, Path, Path]:
    stem = normalize_date_for_filename(date_start, date_end)
    xlsx = Path(output_xlsx) if output_xlsx else EXCEL_DIR / spec["output_patterns"]["normalized_xlsx"].format(date_segment=stem)
    meta = Path(output_meta) if output_meta else ARTIFACTS_DIR / spec["output_patterns"]["meta_json"].format(date_segment=stem)
    native_export = ARTIFACTS_DIR / spec["output_patterns"]["native_export_xlsx"].format(date_segment=stem)
    return xlsx, meta, native_export


def run_auth_refresh() -> dict[str, Any]:
    subprocess.run(["node", str(AUTH_REFRESH_SCRIPT)], cwd=str(REPO_ROOT), check=True)
    return read_json(AUTH_REFRESH_RESULT)


def load_storage_state_cookies(storage_state_path: str) -> list[dict[str, Any]]:
    storage = read_json(Path(storage_state_path))
    return storage["cookies"]


def build_session(cookies: list[dict[str, Any]]) -> requests.Session:
    session = requests.Session()
    for cookie in cookies:
        session.cookies.set(
            cookie["name"],
            cookie["value"],
            domain=cookie.get("domain"),
            path=cookie.get("path", "/"),
        )
    session.headers.update(
        {
            "User-Agent": "DentalPro-RPT3-Builder/1.0",
            "Accept-Language": "ru,en;q=0.9",
        }
    )
    return session


def build_runtime_url(spec: dict[str, Any], date_start: str, date_end: str) -> str:
    base = spec["base_url"].rstrip("/") + spec["route_path"]
    query = urlencode(
        {
            "id": spec["report_id"],
            "date_range[start]": date_start,
            "date_range[end]": date_end,
        }
    )
    return f"{base}?{query}"


def build_export_entry_url(spec: dict[str, Any], date_start: str, date_end: str) -> str:
    payload = {
        "id": str(spec["report_id"]),
        "date_range": {
            "start": date_start,
            "end": date_end,
        },
    }
    encoded = base64.b64encode(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")).decode("ascii")
    params = urlencode(
        {
            "plugin": "reports",
            "schema": "reportsDefault",
            "params[id]": str(spec["report_id"]),
            "params[filter]": encoded,
        }
    )
    return f"{spec['base_url'].rstrip('/')}/exporter/index?{params}"


def is_login_redirect(response: requests.Response, soup: BeautifulSoup) -> bool:
    if "login" in response.url:
        return True
    if soup.select_one('input[name="password"]') and "Войти" in soup.get_text(" ", strip=True):
        return True
    return False


def parse_runtime_table(spec: dict[str, Any], html: str, runtime_url: str) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    expected_headers = spec["runtime_columns"]
    target_table = None
    for table in soup.select("table"):
        headers = [cell.get_text(" ", strip=True) for cell in table.select("thead tr th")]
        if headers[: len(expected_headers)] == expected_headers:
            target_table = table
            break
    if target_table is None:
        raise ValueError("Failed to locate target report table")

    runtime_rows: list[dict[str, Any]] = []
    body_rows = target_table.select("tbody tr")
    for idx, tr in enumerate(body_rows, start=1):
        cells = [td.get_text(" ", strip=True) for td in tr.select("td")]
        if not any(cells):
            continue
        patient_link = tr.select_one('a[href*="/cbase/detail.html?id="]')
        patient_href = patient_link.get("href", "") if patient_link else ""
        patient_id_match = re.search(r"id=(\d+)", patient_href)
        patient_id = int(patient_id_match.group(1)) if patient_id_match else None

        debt = parse_money(cells[2] if len(cells) > 2 else None)
        advance = parse_money(cells[3] if len(cells) > 3 else None)
        available = parse_money(cells[4] if len(cells) > 4 else None)
        last_op = parse_ui_date(cells[5] if len(cells) > 5 else None)

        runtime_rows.append(
            {
                "report_code": spec["report_code"],
                "report_name": spec["report_name"],
                "date_start": None,  # filled later
                "date_end": None,  # filled later
                "row_source": "runtime_html",
                "row_index": idx,
                "patient_name": cells[0] if len(cells) > 0 else None,
                "phone": normalize_phone(cells[1] if len(cells) > 1 else None),
                "debt": debt,
                "advance": advance,
                "available_own_funds": available,
                "last_operation_date": last_op,
                "arithmetic_ok": bool_close(available, (advance or 0.0) - (debt or 0.0)),
                "patient_id": patient_id,
                "identity_proof": "row_link_patient_card" if patient_id is not None else "not_proven",
                "debt_source_status": "sample_proven_cashbox_based_family_only",
                "runtime_export_parity_status": "not_checked",
                "notes": "",
                "_patient_href": urljoin(runtime_url, patient_href) if patient_href else None,
            }
        )

    footer_cells = [cell.get_text(" ", strip=True) for cell in target_table.select("tfoot tr td, tfoot tr th")]
    footer = {
        "debt": parse_money(footer_cells[2]) if len(footer_cells) > 2 else None,
        "advance": parse_money(footer_cells[3]) if len(footer_cells) > 3 else None,
        "available_own_funds": parse_money(footer_cells[4]) if len(footer_cells) > 4 else None,
    }

    export_link = None
    for anchor in soup.select("a[href]"):
        href = anchor.get("href", "")
        if "/exporter/index?plugin=reports" in href:
            export_link = urljoin(runtime_url, href)
            break

    return {
        "soup": soup,
        "runtime_rows": runtime_rows,
        "runtime_columns": expected_headers,
        "runtime_row_count": len(runtime_rows),
        "runtime_footer": footer,
        "export_link": export_link,
        "title": soup.title.get_text(strip=True) if soup.title else "",
    }


def fetch_native_export(
    session: requests.Session,
    spec: dict[str, Any],
    date_start: str,
    date_end: str,
    native_export_path: Path,
    problem_log: list[dict[str, Any]],
) -> dict[str, Any]:
    entry_url = build_export_entry_url(spec, date_start, date_end)
    result: dict[str, Any] = {
        "checked": True,
        "available": False,
        "flow_type": "two_step_async_export",
        "file_path": None,
        "row_count": None,
        "footer": {
            "debt": None,
            "advance": None,
            "available_own_funds": None,
        },
        "entry_url": entry_url,
        "process_export_url": None,
        "download_url": None,
    }

    entry_resp = session.get(entry_url, timeout=120)
    entry_resp.raise_for_status()

    if entry_resp.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        native_export_path.parent.mkdir(parents=True, exist_ok=True)
        native_export_path.write_bytes(entry_resp.content)
        result["available"] = True
        result["file_path"] = relative_repo_path(native_export_path)
        return result

    entry_soup = BeautifulSoup(entry_resp.text, "html.parser")
    process_url = None
    step_url = entry_url.replace("/exporter/index?", "/exporter/step_export?")
    step_resp = session.get(step_url, timeout=120)
    step_resp.raise_for_status()
    step_soup = BeautifulSoup(step_resp.text, "html.parser")
    iframe = step_soup.select_one("iframe[src]")
    if iframe:
        process_url = urljoin(spec["base_url"], iframe.get("src"))
        result["process_export_url"] = process_url

    if not process_url:
        problem_log.append(
            {
                "code": "EXPORT_FLOW_PARSE_FAILED",
                "severity": "high",
                "message": "Failed to resolve process_export iframe for native export",
            }
        )
        return result

    process_resp = session.get(process_url, timeout=120)
    process_resp.raise_for_status()
    match = re.search(r"export_done\('([^']+\.xlsx)'\)", process_resp.text)
    if not match:
        problem_log.append(
            {
                "code": "EXPORT_DOWNLOAD_PATH_MISSING",
                "severity": "high",
                "message": "process_export did not return export_done(file.xlsx)",
            }
        )
        return result

    download_url = urljoin(spec["base_url"], match.group(1))
    result["download_url"] = download_url
    file_resp = session.get(download_url, timeout=120)
    file_resp.raise_for_status()
    native_export_path.parent.mkdir(parents=True, exist_ok=True)
    native_export_path.write_bytes(file_resp.content)
    result["available"] = True
    result["file_path"] = relative_repo_path(native_export_path)
    return result


def parse_export_workbook(spec: dict[str, Any], path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for idx, row in enumerate(rows):
        if tuple(row[:6]) == tuple(spec["runtime_columns"]):
            header_idx = idx
            break
    if header_idx is None:
        raise ValueError("Failed to find export header row")

    export_rows: list[dict[str, Any]] = []
    footer = {
        "debt": None,
        "advance": None,
        "available_own_funds": None,
    }

    for row in rows[header_idx + 1 :]:
        values = row[:6]
        if not any(v is not None for v in values):
            continue
        if str(values[0]).strip() == "ИТОГО":
            footer = {
                "debt": parse_money(values[2]),
                "advance": parse_money(values[3]),
                "available_own_funds": parse_money(values[4]),
            }
            continue
        last_op = values[5]
        if isinstance(last_op, datetime):
            last_op = last_op.date()
        export_rows.append(
            {
                "patient_name": str(values[0]) if values[0] is not None else None,
                "phone": normalize_phone(str(values[1]) if values[1] is not None else None),
                "debt": parse_money(values[2]),
                "advance": parse_money(values[3]),
                "available_own_funds": parse_money(values[4]),
                "last_operation_date": last_op if isinstance(last_op, date) else None,
            }
        )

    return {
        "row_count": len(export_rows),
        "footer": footer,
        "rows": export_rows,
    }


def row_key(row: dict[str, Any]) -> str:
    return f"{(row.get('patient_name') or '').strip()}|{normalize_phone(row.get('phone')) or ''}"


def iso_or_none(value: date | None) -> str | None:
    return value.isoformat() if isinstance(value, date) else None


def compare_runtime_export(
    runtime_rows: list[dict[str, Any]],
    runtime_footer: dict[str, float | None],
    export_rows: list[dict[str, Any]],
    export_footer: dict[str, float | None],
) -> dict[str, Any]:
    runtime_map = {row_key(row): row for row in runtime_rows}
    export_map = {row_key(row): row for row in export_rows}

    runtime_only = [runtime_map[key] for key in sorted(set(runtime_map) - set(export_map))]
    export_only = [export_map[key] for key in sorted(set(export_map) - set(runtime_map))]
    field_mismatches: list[dict[str, Any]] = []

    for key in sorted(set(runtime_map) & set(export_map)):
        runtime_row = runtime_map[key]
        export_row = export_map[key]
        diffs = {}
        for field in ("debt", "advance", "available_own_funds"):
            if not bool_close(runtime_row.get(field), export_row.get(field)):
                diffs[field] = {
                    "runtime": runtime_row.get(field),
                    "export": export_row.get(field),
                }
        runtime_last = iso_or_none(runtime_row.get("last_operation_date"))
        export_last = iso_or_none(export_row.get("last_operation_date"))
        if runtime_last != export_last:
            diffs["last_operation_date"] = {
                "runtime": runtime_last,
                "export": export_last,
            }
        if diffs:
            field_mismatches.append(
                {
                    "patient_name": runtime_row.get("patient_name"),
                    "phone": runtime_row.get("phone"),
                    "diffs": diffs,
                }
            )

    footer_mismatch = {
        metric: {
            "runtime": runtime_footer.get(metric),
            "export": export_footer.get(metric),
        }
        for metric in ("debt", "advance", "available_own_funds")
        if not bool_close(runtime_footer.get(metric), export_footer.get(metric))
    }

    mismatch_detected = bool(runtime_only or export_only or field_mismatches or footer_mismatch)
    return {
        "status": "mismatch_detected" if mismatch_detected else "matched",
        "mismatch_detected": mismatch_detected,
        "runtime_row_count": len(runtime_rows),
        "export_row_count": len(export_rows),
        "runtime_only_count": len(runtime_only),
        "runtime_only_rows": [
            {"patient_name": row.get("patient_name"), "phone": row.get("phone")}
            for row in runtime_only
        ],
        "export_only_count": len(export_only),
        "export_only_rows": [
            {"patient_name": row.get("patient_name"), "phone": row.get("phone")}
            for row in export_only
        ],
        "field_mismatch_count": len(field_mismatches),
        "field_mismatches": field_mismatches,
        "footer_mismatch": footer_mismatch,
        "notes": [],
    }


def apply_row_parity_notes(runtime_rows: list[dict[str, Any]], parity: dict[str, Any]) -> None:
    runtime_only_keys = {
        f"{row['patient_name']}|{normalize_phone(row.get('phone')) or ''}"
        for row in parity.get("runtime_only_rows", [])
    }
    mismatch_map = {
        f"{row['patient_name']}|{normalize_phone(row.get('phone')) or ''}": row["diffs"]
        for row in parity.get("field_mismatches", [])
    }

    for row in runtime_rows:
        key = row_key(row)
        if parity["status"] == "not_checked":
            row["runtime_export_parity_status"] = "not_checked"
        elif key in runtime_only_keys:
            row["runtime_export_parity_status"] = "runtime_only"
            row["notes"] = "missing_in_native_export"
        elif key in mismatch_map:
            row["runtime_export_parity_status"] = "field_mismatch"
            row["notes"] = f"export_diff_fields={','.join(sorted(mismatch_map[key].keys()))}"
        else:
            row["runtime_export_parity_status"] = "row_match"


def build_summary_row(
    spec: dict[str, Any],
    date_start: str,
    date_end: str,
    runtime_rows: list[dict[str, Any]],
    runtime_footer: dict[str, float | None],
    export_result: dict[str, Any],
    parity: dict[str, Any],
    generated_at: str,
) -> dict[str, Any]:
    export_footer = export_result.get("footer", {})
    runtime_arithmetic_ok = bool_close(
        runtime_footer.get("available_own_funds"),
        (runtime_footer.get("advance") or 0.0) - (runtime_footer.get("debt") or 0.0),
    )
    return {
        "report_code": spec["report_code"],
        "report_name": spec["report_name"],
        "date_start": date_start,
        "date_end": date_end,
        "runtime_row_count": len(runtime_rows),
        "runtime_footer_debt": runtime_footer.get("debt"),
        "runtime_footer_advance": runtime_footer.get("advance"),
        "runtime_footer_available_own_funds": runtime_footer.get("available_own_funds"),
        "runtime_arithmetic_ok": runtime_arithmetic_ok,
        "export_checked": export_result.get("checked", False),
        "export_row_count": export_result.get("row_count"),
        "export_footer_debt": export_footer.get("debt"),
        "export_footer_advance": export_footer.get("advance"),
        "export_footer_available_own_funds": export_footer.get("available_own_funds"),
        "mismatch_detected": parity.get("mismatch_detected", False),
        "safe_operational_status": "ready_with_limits",
        "generated_at": generated_at,
    }


def build_excel(
    workbook_path: Path,
    normalized_rows: list[dict[str, Any]],
    summary_row: dict[str, Any],
    warnings: list[dict[str, Any]],
) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_norm = wb.active
    ws_norm.title = "RPT_3_normalized"
    write_table(ws_norm, 1, normalized_rows)

    ws_summary = wb.create_sheet("RPT_3_summary")
    write_table(ws_summary, 1, [summary_row])

    ws_warn = wb.create_sheet("RPT_3_warnings")
    write_table(ws_warn, 1, warnings)
    for cell in ws_warn[2]:
        cell.fill = WARN_FILL

    for ws in wb.worksheets:
        sheet_autofit(ws)
    wb.save(workbook_path)


def main() -> None:
    load_dotenv(REPO_ROOT / ".env", override=False)
    spec = read_json(SPEC_PATH)
    cli = parse_cli(spec)

    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--no-export", action="store_true")
    parser.add_argument("--skip-auth-refresh", action="store_true")
    parser.add_argument("--output-xlsx")
    parser.add_argument("--output-meta")
    known, _ = parser.parse_known_args()

    output_xlsx, output_meta, native_export_path = get_output_paths(
        spec, cli.date_start, cli.date_end, known.output_xlsx, known.output_meta
    )

    warnings: list[dict[str, Any]] = [
        make_warning(
            "IBALANCE_NOT_PROVEN",
            "medium",
            "source_mapping",
            "Do not promote i/balance to direct read source for RPT_3.",
        ),
        make_warning(
            "DEBT_SOURCE_SAMPLE_ONLY",
            "medium",
            "source_mapping",
            "Debt source is sample-proven via patient cashbox balances, not fully backend-closed.",
        ),
    ]

    meta: dict[str, Any] = {
        "report_code": spec["report_code"],
        "report_name": spec["report_name"],
        "date_start": cli.date_start,
        "date_end": cli.date_end,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "run_status": "started",
        "auth": {
            "authenticated": False,
            "auth_mode": "auth_refresh_js_plus_storage_state",
            "redirect_detected": False,
        },
        "runtime": {
            "url": None,
            "columns": [],
            "row_count": 0,
            "footer": {
                "debt": None,
                "advance": None,
                "available_own_funds": None,
            },
        },
        "export": {
            "checked": False,
            "available": False,
            "flow_type": spec["export_flow_type"],
            "file_path": None,
            "row_count": None,
            "footer": {
                "debt": None,
                "advance": None,
                "available_own_funds": None,
            },
        },
        "parity": {
            "status": "not_checked",
            "mismatch_detected": False,
            "notes": [],
        },
        "source_classification": {
            "primary_truth": "runtime_html",
            "secondary_truth": "native_export_optional",
            "debt_source_status": "sample_proven_cashbox_based",
            "ibalance_status": "not_proven_as_direct_read_source",
        },
        "warnings": warnings,
        "problem_log": [],
    }

    try:
        if not known.skip_auth_refresh:
            auth_result = run_auth_refresh()
        else:
            auth_result = read_json(AUTH_REFRESH_RESULT)
        meta["auth"]["authenticated"] = bool(auth_result.get("authenticated"))
        meta["auth"]["refresh_result_path"] = relative_repo_path(AUTH_REFRESH_RESULT)
        if not auth_result.get("authenticated"):
            meta["auth"]["redirect_detected"] = True
            warnings.append(
                make_warning(
                    "AUTH_REDIRECT_DETECTED",
                    "high",
                    "auth",
                    "Auth refresh did not reach an authenticated state.",
                )
            )
            meta["run_status"] = "auth_failed"
            write_json(output_meta, meta)
            raise SystemExit("Auth refresh failed")

        session = build_session(load_storage_state_cookies(auth_result["storage_state_path"]))
        runtime_url = build_runtime_url(spec, cli.date_start, cli.date_end)
        runtime_response = session.get(runtime_url, timeout=120)
        runtime_response.raise_for_status()
        runtime_soup = BeautifulSoup(runtime_response.text, "html.parser")
        if is_login_redirect(runtime_response, runtime_soup):
            meta["auth"]["redirect_detected"] = True
            warnings.append(
                make_warning(
                    "AUTH_REDIRECT_DETECTED",
                    "high",
                    "runtime",
                    "Runtime report request redirected to login or rendered login form.",
                )
            )
            meta["run_status"] = "auth_redirect_detected"
            write_json(output_meta, meta)
            raise SystemExit("Runtime report request redirected to login")

        parsed_runtime = parse_runtime_table(spec, runtime_response.text, runtime_url)
        normalized_rows = parsed_runtime["runtime_rows"]
        for row in normalized_rows:
            row["date_start"] = cli.date_start
            row["date_end"] = cli.date_end
        meta["runtime"] = {
            "url": runtime_url,
            "columns": parsed_runtime["runtime_columns"],
            "row_count": parsed_runtime["runtime_row_count"],
            "footer": parsed_runtime["runtime_footer"],
        }

        export_result = copy.deepcopy(meta["export"])
        parity = copy.deepcopy(meta["parity"])
        if known.no_export:
            warnings.append(
                make_warning(
                    "EXPORT_NOT_CHECKED",
                    "low",
                    "export",
                    "Native export pass was intentionally skipped.",
                )
            )
        else:
            export_result = fetch_native_export(
                session,
                spec,
                cli.date_start,
                cli.date_end,
                native_export_path,
                meta["problem_log"],
            )
            if export_result.get("available"):
                parsed_export = parse_export_workbook(spec, native_export_path)
                export_result["row_count"] = parsed_export["row_count"]
                export_result["footer"] = parsed_export["footer"]
                parity = compare_runtime_export(
                    normalized_rows,
                    parsed_runtime["runtime_footer"],
                    parsed_export["rows"],
                    parsed_export["footer"],
                )
                if parity["mismatch_detected"]:
                    warnings.append(
                        make_warning(
                            "EXPORT_RUNTIME_MISMATCH",
                            "high",
                            "parity",
                            "Native export and runtime HTML differ for this run; runtime remains primary truth.",
                        )
                    )
                    meta["problem_log"].append(
                        {
                            "code": "EXPORT_RUNTIME_MISMATCH",
                            "severity": "high",
                            "details": {
                                "runtime_row_count": parity["runtime_row_count"],
                                "export_row_count": parity["export_row_count"],
                                "runtime_only_count": parity["runtime_only_count"],
                                "export_only_count": parity["export_only_count"],
                                "field_mismatch_count": parity["field_mismatch_count"],
                                "footer_mismatch": parity["footer_mismatch"],
                            },
                        }
                    )
            else:
                warnings.append(
                    make_warning(
                        "EXPORT_NOT_AVAILABLE",
                        "medium",
                        "export",
                        "Native export did not resolve to a downloadable xlsx in this run.",
                    )
                )

        apply_row_parity_notes(normalized_rows, parity)

        excel_rows = []
        for row in normalized_rows:
            excel_rows.append(
                {
                    "report_code": row["report_code"],
                    "report_name": row["report_name"],
                    "date_start": row["date_start"],
                    "date_end": row["date_end"],
                    "row_source": row["row_source"],
                    "row_index": row["row_index"],
                    "patient_name": row["patient_name"],
                    "phone": row["phone"],
                    "debt": row["debt"],
                    "advance": row["advance"],
                    "available_own_funds": row["available_own_funds"],
                    "last_operation_date": row["last_operation_date"],
                    "arithmetic_ok": row["arithmetic_ok"],
                    "patient_id": row["patient_id"],
                    "identity_proof": row["identity_proof"],
                    "debt_source_status": row["debt_source_status"],
                    "runtime_export_parity_status": row["runtime_export_parity_status"],
                    "notes": row["notes"],
                }
            )

        summary_row = build_summary_row(
            spec,
            cli.date_start,
            cli.date_end,
            normalized_rows,
            parsed_runtime["runtime_footer"],
            export_result,
            parity,
            meta["generated_at"],
        )

        build_excel(output_xlsx, excel_rows, summary_row, warnings)

        meta["export"] = export_result
        meta["parity"] = parity
        meta["warnings"] = warnings
        meta["run_status"] = "success_with_warnings" if any(w["severity"] in {"medium", "high"} for w in warnings) else "success"
        write_json(output_meta, meta)

        print(json.dumps(
            {
                "xlsx": str(output_xlsx),
                "meta": str(output_meta),
                "native_export": str(native_export_path) if export_result.get("available") else None,
                "run_status": meta["run_status"],
            },
            ensure_ascii=False,
        ))
    except Exception as exc:
        meta["run_status"] = "error"
        meta["problem_log"].append(
            {
                "code": "BUILDER_ERROR",
                "severity": "critical",
                "details": str(exc),
            }
        )
        write_json(output_meta, meta)
        raise


if __name__ == "__main__":
    main()
