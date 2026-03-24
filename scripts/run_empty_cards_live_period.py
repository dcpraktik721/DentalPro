#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlencode, urljoin, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from rpt50_card_info_normalizer import WARNING_CODE, build_card_info


ROOT = Path("/Users/macbook15/Downloads/MacAi/DentalPro")
EXCEL_DIR = ROOT / "excel"
ARTIFACTS_DIR = ROOT / "artifacts"
REPORTS_DIR = ROOT / "reports"
STORAGE_PATH = Path(
    "/Users/macbook15/Downloads/YDIREKT_code/mac-ai-os/dentalpro-playwright/auth/dentalpro.storage.json"
)
BASE_URL = "https://dcpraktik.dental-pro.online/medblock/cards/index"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run live empty-cards extraction for a period.")
    parser.add_argument("--start", required=True, help="YYYY-MM-DD")
    parser.add_argument("--end", required=True, help="YYYY-MM-DD")
    parser.add_argument("--compare-workbook", help="Previous workbook path for comparison")
    parser.add_argument("--compare-result-json", help="Previous result JSON path for comparison")
    return parser.parse_args()


def load_storage_cookies() -> list[dict[str, Any]]:
    data = json.loads(STORAGE_PATH.read_text())
    return data.get("cookies", [])


def build_session() -> requests.Session:
    session = requests.Session()
    for cookie in load_storage_cookies():
        session.cookies.set(
            cookie["name"],
            cookie["value"],
            domain=cookie.get("domain"),
            path=cookie.get("path", "/"),
        )
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0",
            "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        }
    )
    return session


def build_filtered_url(date_start: str, date_end: str, page: int | None = None) -> str:
    params: list[tuple[str, str]] = [
        ("filter[Filter][status_change_date][start]", date_start),
        ("filter[Filter][status_change_date][end]", date_end),
        ("filter[Filter][status][]", "0"),
        ("filter[Filter][client_id]", ""),
        ("filter[Filter][diagnosis_id]", ""),
        ("filter[Filter][appointment_id]", ""),
    ]
    if page is not None:
        params.append(("page", str(page)))
    return f"{BASE_URL}?{urlencode(params)}"


def parse_datetime_from_created(created_raw: str) -> str | None:
    match = re.search(r"(\d{2}\.\d{2}\.\d{4})\s+в\s+(\d{2}:\d{2}:\d{2})", created_raw or "")
    if not match:
        return None
    return datetime.strptime(f"{match.group(1)} {match.group(2)}", "%d.%m.%Y %H:%M:%S").isoformat()


def parse_patient_id(patient_href: str) -> str | None:
    if not patient_href:
        return None
    parsed = urlparse(patient_href)
    return parse_qs(parsed.query).get("id", [None])[0]


def parse_card_id_from_href(detail_href: str) -> str | None:
    if not detail_href:
        return None
    parsed = urlparse(detail_href)
    return parse_qs(parsed.query).get("id", [None])[0]


def normalize_status(status_raw: str) -> str:
    status = " ".join((status_raw or "").split()).strip()
    if status == "Пустая":
        return "Незаполненная"
    return status


def parse_page(html: str, url: str, page_label: str) -> dict[str, Any]:
    soup = BeautifulSoup(html, "html.parser")
    title = soup.title.get_text(" ", strip=True) if soup.title else ""
    headers = [th.get_text(" ", strip=True) for th in soup.select("table thead th")]

    rows: list[dict[str, Any]] = []
    for idx, tr in enumerate(soup.select("table tbody tr"), start=1):
        tds = tr.select("td")
        if len(tds) < 6:
            continue
        patient_link = tr.select_one('a[href*="cbase/detail.html?id="]')
        detail_link = tr.select_one('a[href*="/medblock/cards/view?id="]')
        info_raw = tds[1].get_text(" ", strip=True)
        info_norm = build_card_info(primary_text=info_raw, status_text=tds[4].get_text(" ", strip=True))
        created_raw = tds[5].get_text(" ", strip=True)
        row = {
            "source_page": page_label,
            "source_url": url,
            "row_index": idx,
            "card_id": tds[0].get_text(" ", strip=True),
            "detail_card_id": parse_card_id_from_href(detail_link.get("href", "") if detail_link else ""),
            "patient_name": tds[2].get_text(" ", strip=True),
            "patient_id": parse_patient_id(patient_link.get("href", "") if patient_link else ""),
            "patient_href": patient_link.get("href", "") if patient_link else "",
            "doctor_name": tds[3].get_text(" ", strip=True),
            "status_raw": tds[4].get_text(" ", strip=True),
            "status_normalized": normalize_status(tds[4].get_text(" ", strip=True)),
            "info_raw": info_raw,
            "info_normalized": info_norm.value,
            "info_warnings": info_norm.warnings,
            "created_raw": created_raw,
            "created_at": parse_datetime_from_created(created_raw),
            "updated_raw": tds[6].get_text(" ", strip=True) if len(tds) > 6 else "",
            "detail_href": detail_link.get("href", "") if detail_link else "",
        }
        rows.append(row)

    page_links = []
    for a in soup.select(".pagination a, ul.pagination a"):
        href = a.get("href")
        if href:
            page_links.append(urljoin(url, href))
    page_links = list(dict.fromkeys(page_links))

    controls_text = " ".join(el.get_text(" ", strip=True) for el in soup.select("a,button"))
    result = {
        "url": url,
        "title": title,
        "headers": headers,
        "row_count": len(rows),
        "rows": rows,
        "page_links": page_links,
        "export_detected": "Экспорт" in controls_text,
        "empty_state_detected": "Нет данных" in soup.get_text(" ", strip=True),
        "date_controls_present": bool(soup.select_one('input[name="filter[Filter][status_change_date][start]"]'))
        and bool(soup.select_one('input[name="filter[Filter][status_change_date][end]"]')),
        "status_filter_present": bool(soup.select_one('select[name="filter[Filter][status][]"]')),
        "doctor_filter_present": bool(soup.select_one('select[name*="doctor"]')),
    }
    return result


def dedupe_rows(page_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    deduped: list[dict[str, Any]] = []
    seen: set[str] = set()
    duplicates: list[str] = []
    for row in page_rows:
        key = str(row.get("card_id") or row.get("detail_card_id") or "")
        if not key:
            continue
        if key in seen:
            duplicates.append(key)
            continue
        seen.add(key)
        deduped.append(row)
    return deduped, duplicates


def get_previous_comparison(previous_result_json: str | None, previous_workbook: str | None) -> dict[str, Any]:
    comparison: dict[str, Any] = {}
    if previous_result_json:
        data = json.loads(Path(previous_result_json).read_text())
        comparison["previous_claimed_records_count"] = data.get("records_count")
        comparison["previous_run_status"] = data.get("run_status")
    if previous_workbook:
        wb = openpyxl.load_workbook(previous_workbook, data_only=True)
        ws = wb["Незаполненные карты"]
        numeric_card_ids = 0
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row_idx, 1).value
            if isinstance(value, str) and re.fullmatch(r"\d+", value.strip()):
                numeric_card_ids += 1
        comparison["previous_workbook_path"] = previous_workbook
        comparison["previous_workbook_max_row_minus_header"] = ws.max_row - 1
        comparison["previous_workbook_numeric_card_ids"] = numeric_card_ids
        comparison["previous_workbook_contaminated_main_sheet"] = (ws.max_row - 1) != numeric_card_ids
    return comparison


def write_workbook(path: Path, rows: list[dict[str, Any]], source_meta: dict[str, Any], issues: list[dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    main_ws = wb.active
    main_ws.title = "Незаполненные карты"
    main_ws.append(["ID карты", "Врач", "Пациент", "Информация о карте", "Статус"])
    for cell in main_ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        main_ws.append(
            [
                row["card_id"],
                row["doctor_name"],
                row["patient_name"],
                row["info_normalized"],
                row["status_normalized"],
            ]
        )
    main_ws.freeze_panes = "A2"
    main_ws.auto_filter.ref = main_ws.dimensions

    doctor_counts = sorted(Counter(row["doctor_name"] for row in rows if row["doctor_name"]).items(), key=lambda x: (-x[1], x[0]))

    sum_ws = wb.create_sheet("Свод")
    summary_rows = [
        ("Незаполненные карты", None),
        ("Период", f"{source_meta['date_from']} -> {source_meta['date_to']}"),
        ("Сформировано", source_meta["generated_at"]),
        ("Всего карт", len(rows)),
        ("Всего врачей", len(doctor_counts)),
        ("Страниц собрано", source_meta["pages_collected"]),
        ("Источник истины", "browser-authenticated direct GET runtime HTML"),
        ("Алгоритм", "full live rerun across all visible pagination pages; no baseline merge"),
    ]
    for item in summary_rows:
        sum_ws.append(item)
    sum_ws.append([None, None])
    sum_ws.append(["Issues", "Detail"])
    for issue in issues:
        sum_ws.append([issue["code"], issue["message"]])
    sum_ws.append([None, None])
    sum_ws.append(["Врач", "Количество незаполненных карт"])
    for doctor, count in doctor_counts:
        sum_ws.append([doctor, count])
    for cell in sum_ws[1]:
        cell.font = Font(bold=True)
    for cell in sum_ws[10]:
        cell.font = Font(bold=True)
    for cell in sum_ws[13]:
        cell.font = Font(bold=True)

    src_ws = wb.create_sheet("Источник")
    src_ws.append(["key", "value"])
    for cell in src_ws[1]:
        cell.font = Font(bold=True)
    for key, value in source_meta.items():
        if isinstance(value, (list, dict)):
            value = json.dumps(value, ensure_ascii=False)
        src_ws.append([key, value])

    doctor_ws = wb.create_sheet("По врачам")
    doctor_ws.append(["Врач", "Количество незаполненных карт"])
    for cell in doctor_ws[1]:
        cell.font = Font(bold=True)
    for doctor, count in doctor_counts:
        doctor_ws.append([doctor, count])
    doctor_ws.freeze_panes = "A2"
    doctor_ws.auto_filter.ref = doctor_ws.dimensions

    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            width = max(len(str(cell.value or "")) for cell in ws[get_column_letter(col)])
            ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 14), 80)

    wb.save(path)


def write_report(path: Path, result: dict[str, Any]) -> None:
    lines = [
        "# Ревизия алгоритма пустых карт - live rerun",
        "",
        "## Summary",
        f"- период: `{result['date_from']} -> {result['date_to']}`",
        f"- run_status: `{result['run_status']}`",
        f"- route_status: `{result['route_status']}`",
        f"- records_count: `{result['records_count']}`",
        f"- pages_collected: `{result['pages_collected']}`",
        f"- source_of_truth: `browser-authenticated direct GET runtime HTML`",
        "",
        "## Revised Method",
        "- exact filtered route `/medblock/cards/index`",
        "- exact GET filter contract with `status_change_date[start/end]` and `status[]=0`",
        "- full fresh rerun across all visible pagination pages",
        "- no baseline workbook merge",
        "- no export substitution",
        "",
        "## Issues",
    ]
    for issue in result["issues"]:
        lines.append(f"- `{issue['code']}` - {issue['message']}")
    lines.extend(["", "## Comparison With Previous Algorithm"])
    cmp = result.get("comparison", {})
    if cmp:
        for key, value in cmp.items():
            lines.append(f"- `{key}` - `{value}`")
    else:
        lines.append("- comparison data not available")
    lines.extend(
        [
            "",
            "## Claim Boundary",
            result["claim_boundary"],
            "",
            "## Files",
        ]
    )
    for file_path in result["created_files"]:
        lines.append(f"- `{file_path}`")
    path.write_text("\n".join(lines) + "\n")


def main() -> int:
    args = parse_args()
    session = build_session()
    issues: list[dict[str, str]] = []

    base_url = build_filtered_url(args.start, args.end)
    first_resp = session.get(base_url, timeout=60)
    route_status = "opened_authenticated"
    if "/user/login" in first_resp.url:
        route_status = "login_redirect"
        issues.append(
            {
                "code": "AUTH_REDIRECT_DETECTED",
                "severity": "high",
                "message": "Authenticated direct GET fell back to login; extraction not trusted.",
            }
        )

    first_page = parse_page(first_resp.text, first_resp.url, "page_0")
    normalized_headers = [header for header in first_page["headers"] if header]
    headers_ok = normalized_headers == ["ID", "Информация о карте", "Пациент", "Врач", "Статус", "Создано", "Обновлено"]
    if not headers_ok:
        issues.append(
            {
                "code": "SCHEMA_DRIFT",
                "severity": "high",
                "message": f"Unexpected headers: {first_page['headers']}",
            }
        )
    if not first_page["export_detected"]:
        issues.append(
            {
                "code": "EXPORT_NOT_DETECTED",
                "severity": "low",
                "message": "Native export control was not detected in this live rerun.",
            }
        )

    page_urls = [first_resp.url] + [u for u in first_page["page_links"] if u != first_resp.url]
    all_page_results = [first_page]
    for page_index, url in enumerate(page_urls[1:], start=1):
        resp = session.get(url, timeout=60)
        all_page_results.append(parse_page(resp.text, resp.url, f"page_{page_index}"))

    all_rows = [row for page in all_page_results for row in page["rows"]]
    rows, duplicates = dedupe_rows(all_rows)
    if duplicates:
        issues.append(
            {
                "code": "DUPLICATE_CARD_IDS",
                "severity": "medium",
                "message": f"Duplicate card ids were seen across pages: {len(duplicates)}",
            }
        )
    if not rows:
        issues.append(
            {
                "code": "EMPTY_RESULT_SET",
                "severity": "medium",
                "message": "Filtered route returned zero data rows.",
            }
        )

    comparison = get_previous_comparison(args.compare_result_json, args.compare_workbook)
    if "previous_claimed_records_count" in comparison and comparison["previous_claimed_records_count"] != len(rows):
        issues.append(
            {
                "code": "COUNT_DRIFT_VS_PREVIOUS_RESULT",
                "severity": "medium",
                "message": f"Previous claimed count {comparison['previous_claimed_records_count']} vs current live count {len(rows)}.",
            }
        )
    if comparison.get("previous_workbook_contaminated_main_sheet"):
        issues.append(
            {
                "code": "PREVIOUS_WORKBOOK_MAIN_SHEET_CONTAMINATED",
                "severity": "medium",
                "message": "Previous workbook main sheet contained non-card rows beyond numeric card ids.",
            }
        )

    timestamp = datetime.now().isoformat(timespec="seconds")
    excel_path = EXCEL_DIR / "Незаполненные карты 2025-01-01 - 2026-03-23 ревизия.xlsx"
    json_path = ARTIFACTS_DIR / "nezapolnennye_karty_2025-01-01_2026-03-23_reaudit_result.json"
    report_path = REPORTS_DIR / "nezapolnennye_karty_2025-01-01_2026-03-23_reaudit_report.md"

    source_meta = {
        "report_code": "RPT_50 / direct_route",
        "date_from": args.start,
        "date_to": args.end,
        "generated_at": timestamp,
        "entry_url": BASE_URL,
        "final_urls": [page["url"] for page in all_page_results],
        "pages_collected": len(all_page_results),
        "headers": first_page["headers"],
        "route_mode": "authenticated direct GET",
        "filter_contract": {
            "filter[Filter][status_change_date][start]": args.start,
            "filter[Filter][status_change_date][end]": args.end,
            "filter[Filter][status][]": "0",
        },
    }

    write_workbook(excel_path, rows, source_meta, issues)

    result = {
        "report_code": "RPT_50",
        "report_title": "Карты пациентов / Незаполненные карты",
        "run_mode": "full_live_rerun_reaudit",
        "date_from": args.start,
        "date_to": args.end,
        "generated_at": timestamp,
        "run_status": "success_with_warnings" if issues else "success",
        "route_status": route_status,
        "records_count": len(rows),
        "pages_collected": len(all_page_results),
        "source_of_truth": "browser-authenticated direct GET runtime HTML",
        "page_stats": [
            {"url": page["url"], "row_count": page["row_count"], "page_label": page["rows"][0]["source_page"] if page["rows"] else page["url"]}
            for page in all_page_results
        ],
        "issues": issues,
        "comparison": comparison,
        "claim_boundary": (
            "Ревизованный алгоритм подтверждает exact filtered direct route /medblock/cards/index, "
            "browser-authenticated direct GET, сбор строк со всех видимых страниц пагинации и нормализацию "
            "поля 'Информация о карте' без baseline merge. Export layer не доказан. Полнота вне текущей "
            "пагинации runtime response не заявляется."
        ),
        "records": rows,
        "created_files": [str(excel_path), str(json_path), str(report_path)],
    }

    json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2))
    write_report(report_path, result)
    print(json.dumps({"records_count": len(rows), "pages_collected": len(all_page_results), "files": result["created_files"]}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
